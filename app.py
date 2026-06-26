# -*- coding: utf-8 -*-
"""
CongoSchool - Application de Gestion Scolaire
Flask + SQLite | Tout en un seul fichier
"""

import sqlite3
import os
from datetime import datetime, date
from flask import Flask, request, redirect, url_for, render_template_string, flash, jsonify, send_file, session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import zipfile
import shutil
from io import BytesIO
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'congoschool-v2-secure-a7f3e9b1d4c6')

# Vercel = in-memory DB; local = file DB
_IS_VERCEL = os.environ.get('VERCEL', '') == '1'
DB_PATH = '/tmp/congoschool.db' if _IS_VERCEL else os.path.join(os.path.dirname(os.path.abspath(__file__)), 'congoschool.db')


# ──────────────────────────────────────────────
# DATABASE
# ──────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = get_db()
    c = conn.cursor()

    c.executescript('''
        CREATE TABLE IF NOT EXISTS classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            level TEXT,
            capacity INTEGER DEFAULT 40,
            teacher_id INTEGER,
            description TEXT,
            cycle TEXT DEFAULT 'primaire',
            FOREIGN KEY (teacher_id) REFERENCES teachers(id)
        );
        CREATE TABLE IF NOT EXISTS teachers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            phone TEXT,
            email TEXT,
            subject TEXT,
            address TEXT,
            hire_date TEXT,
            status TEXT DEFAULT 'actif',
            salary REAL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            matricule TEXT UNIQUE NOT NULL,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            date_of_birth TEXT,
            gender TEXT,
            class_id INTEGER,
            parent_name TEXT,
            parent_phone TEXT,
            address TEXT,
            email TEXT,
            enrollment_date TEXT,
            status TEXT DEFAULT 'actif',
            cycle TEXT DEFAULT 'primaire',
            FOREIGN KEY (class_id) REFERENCES classes(id)
        );
        CREATE TABLE IF NOT EXISTS subjects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            code TEXT,
            coefficient INTEGER DEFAULT 1,
            teacher_id INTEGER,
            class_id INTEGER,
            description TEXT,
            FOREIGN KEY (teacher_id) REFERENCES teachers(id),
            FOREIGN KEY (class_id) REFERENCES classes(id)
        );
        CREATE TABLE IF NOT EXISTS grades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            subject_id INTEGER NOT NULL,
            term TEXT,
            exam_type TEXT,
            score REAL NOT NULL,
            max_score REAL DEFAULT 20,
            date TEXT,
            comment TEXT,
            FOREIGN KEY (student_id) REFERENCES students(id),
            FOREIGN KEY (subject_id) REFERENCES subjects(id)
        );
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            status TEXT NOT NULL,
            comment TEXT,
            FOREIGN KEY (student_id) REFERENCES students(id)
        );
        CREATE TABLE IF NOT EXISTS fees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            fee_type TEXT,
            amount REAL DEFAULT 0,
            amount_paid REAL DEFAULT 0,
            due_date TEXT,
            payment_date TEXT,
            status TEXT DEFAULT 'impaye',
            receipt_number TEXT,
            comment TEXT,
            FOREIGN KEY (student_id) REFERENCES students(id)
        );
        CREATE TABLE IF NOT EXISTS timetable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_id INTEGER NOT NULL,
            subject_id INTEGER NOT NULL,
            day TEXT,
            time_start TEXT,
            time_end TEXT,
            room TEXT,
            FOREIGN KEY (class_id) REFERENCES classes(id),
            FOREIGN KEY (subject_id) REFERENCES subjects(id)
        );
    ''')

    # Ensure cycle column exists (for existing databases)
    try:
        c.execute("ALTER TABLE classes ADD COLUMN cycle TEXT DEFAULT 'primaire'")
    except:
        pass
    try:
        c.execute("ALTER TABLE students ADD COLUMN cycle TEXT DEFAULT 'primaire'")
    except:
        pass

    # Insert demo data only if students table is empty
    count = c.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    if count == 0:
        # Teachers
        c.execute("INSERT INTO teachers (first_name,last_name,phone,email,subject,hire_date,status,salary) VALUES (?,?,?,?,?,?,?,?)",
                  ("Jean", "Mukendi", "+243810000001", "jmukendi@school.cd", "Mathématiques", "2020-09-01", "actif", 450000))
        c.execute("INSERT INTO teachers (first_name,last_name,phone,email,subject,hire_date,status,salary) VALUES (?,?,?,?,?,?,?,?)",
                  ("Marie", "Nsimba", "+243810000002", "mnsimba@school.cd", "Français", "2019-09-01", "actif", 430000))
        c.execute("INSERT INTO teachers (first_name,last_name,phone,email,subject,hire_date,status,salary) VALUES (?,?,?,?,?,?,?,?)",
                  ("Pierre", "Kabongo", "+243810000003", "pkabongo@school.cd", "Sciences", "2021-09-01", "actif", 440000))

        # Classes
        c.execute("INSERT INTO classes (name,level,capacity,teacher_id,description) VALUES (?,?,?,?,?)",
                  ("6ème A", "6ème", 45, 1, "Classe de 6ème année option A"))
        c.execute("INSERT INTO classes (name,level,capacity,teacher_id,description) VALUES (?,?,?,?,?)",
                  ("5ème B", "5ème", 40, 2, "Classe de 5ème année option B"))
        c.execute("INSERT INTO classes (name,level,capacity,teacher_id,description) VALUES (?,?,?,?,?)",
                  ("4ème C", "4ème", 40, 3, "Classe de 4ème année option C"))

        # Students
        today = date.today().isoformat()
        c.execute("INSERT INTO students (matricule,first_name,last_name,date_of_birth,gender,class_id,parent_name,parent_phone,address,enrollment_date,status) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                  ("2025-001", "Emmanuel", "Kasongo", "2012-03-15", "M", 1, "Papa Kasongo", "+243820000001", "Matonge, Kinshasa", today, "actif"))
        c.execute("INSERT INTO students (matricule,first_name,last_name,date_of_birth,gender,class_id,parent_name,parent_phone,address,enrollment_date,status) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                  ("2025-002", "Grace", "Mbemba", "2012-07-22", "F", 1, "Maman Mbemba", "+243820000002", "Lemba, Kinshasa", today, "actif"))
        c.execute("INSERT INTO students (matricule,first_name,last_name,date_of_birth,gender,class_id,parent_name,parent_phone,address,enrollment_date,status) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                  ("2025-003", "Joel", "Tshimanga", "2013-01-10", "M", 2, "Papa Tshimanga", "+243820000003", "Kalamu, Kinshasa", today, "actif"))
        c.execute("INSERT INTO students (matricule,first_name,last_name,date_of_birth,gender,class_id,parent_name,parent_phone,address,enrollment_date,status) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                  ("2025-004", "Divine", "Lukaku", "2013-05-30", "F", 2, "Maman Lukaku", "+243820000004", "Ngiri-Ngiri, Kinshasa", today, "actif"))
        c.execute("INSERT INTO students (matricule,first_name,last_name,date_of_birth,gender,class_id,parent_name,parent_phone,address,enrollment_date,status) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                  ("2025-005", "Patrick", "Mokobe", "2012-11-05", "M", 3, "Papa Mokobe", "+243820000005", "Bandunduville, Kinshasa", today, "actif"))

        # Subjects
        c.execute("INSERT INTO subjects (name,code,coefficient,teacher_id,class_id,description) VALUES (?,?,?,?,?,?)",
                  ("Mathématiques", "MATH", 4, 1, None, "Cours de mathématiques"))
        c.execute("INSERT INTO subjects (name,code,coefficient,teacher_id,class_id,description) VALUES (?,?,?,?,?,?)",
                  ("Français", "FRA", 4, 2, None, "Cours de langue française"))
        c.execute("INSERT INTO subjects (name,code,coefficient,teacher_id,class_id,description) VALUES (?,?,?,?,?,?)",
                  ("Sciences Naturelles", "SCN", 3, 3, None, "Sciences de la nature"))
        c.execute("INSERT INTO subjects (name,code,coefficient,teacher_id,class_id,description) VALUES (?,?,?,?,?,?)",
                  ("Anglais", "ANG", 2, 2, None, "Cours d'anglais"))
        c.execute("INSERT INTO subjects (name,code,coefficient,teacher_id,class_id,description) VALUES (?,?,?,?,?,?)",
                  ("Géographie", "GEO", 2, 1, None, "Géographie générale"))

        # Grades - sample for each student/subject
        grades_data = [
            (1, 1, "1er Trimestre", "Devoir", 15, 20), (1, 2, "1er Trimestre", "Devoir", 12, 20),
            (1, 3, "1er Trimestre", "Devoir", 17, 20), (1, 4, "1er Trimestre", "Devoir", 14, 20),
            (1, 5, "1er Trimestre", "Devoir", 13, 20),
            (2, 1, "1er Trimestre", "Devoir", 18, 20), (2, 2, "1er Trimestre", "Devoir", 16, 20),
            (2, 3, "1er Trimestre", "Devoir", 14, 20), (2, 4, "1er Trimestre", "Devoir", 19, 20),
            (2, 5, "1er Trimestre", "Devoir", 15, 20),
            (3, 1, "1er Trimestre", "Devoir", 11, 20), (3, 2, "1er Trimestre", "Devoir", 13, 20),
            (3, 3, "1er Trimestre", "Devoir", 10, 20), (3, 4, "1er Trimestre", "Devoir", 12, 20),
            (3, 5, "1er Trimestre", "Devoir", 9, 20),
            (4, 1, "1er Trimestre", "Devoir", 16, 20), (4, 2, "1er Trimestre", "Devoir", 17, 20),
            (4, 3, "1er Trimestre", "Devoir", 15, 20), (4, 4, "1er Trimestre", "Devoir", 18, 20),
            (4, 5, "1er Trimestre", "Devoir", 14, 20),
            (5, 1, "1er Trimestre", "Devoir", 13, 20), (5, 2, "1er Trimestre", "Devoir", 11, 20),
            (5, 3, "1er Trimestre", "Devoir", 14, 20), (5, 4, "1er Trimestre", "Devoir", 10, 20),
            (5, 5, "1er Trimestre", "Devoir", 12, 20),
        ]
        for gd in grades_data:
            c.execute("INSERT INTO grades (student_id,subject_id,term,exam_type,score,max_score,date) VALUES (?,?,?,?,?,?,?)",
                      (gd[0], gd[1], gd[2], gd[3], gd[4], gd[5], today))

        # Attendance
        att_statuses = [("présent", ""), ("présent", ""), ("absent", "Maladie"),
                        ("présent", ""), ("retard", "Embouteillage")]
        for i, (st, cm) in enumerate(att_statuses, 1):
            c.execute("INSERT INTO attendance (student_id,date,status,comment) VALUES (?,?,?,?)",
                      (i, today, st, cm))

        # Timetable
        tt_data = [
            (1, 1, "Lundi", "07:30", "09:00", "Salle A1"),
            (1, 2, "Lundi", "09:15", "10:45", "Salle A1"),
            (1, 3, "Mardi", "07:30", "09:00", "Labo Sciences"),
            (1, 4, "Mardi", "09:15", "10:45", "Salle A1"),
            (1, 5, "Mercredi", "07:30", "09:00", "Salle A1"),
            (2, 1, "Lundi", "07:30", "09:00", "Salle B1"),
            (2, 2, "Lundi", "09:15", "10:45", "Salle B1"),
            (2, 3, "Mardi", "07:30", "09:00", "Labo Sciences"),
            (2, 5, "Mercredi", "07:30", "09:00", "Salle B1"),
            (3, 1, "Lundi", "07:30", "09:00", "Salle C1"),
            (3, 3, "Mardi", "07:30", "09:00", "Labo Sciences"),
            (3, 4, "Mercredi", "07:30", "09:00", "Salle C1"),
        ]
        for td in tt_data:
            c.execute("INSERT INTO timetable (class_id,subject_id,day,time_start,time_end,room) VALUES (?,?,?,?,?,?)", td)

        # Fees
        fee_data = [
            (1, "Inscription", 150000, 150000, "2025-09-15", "2025-09-10", "payé", "REC-001"),
            (2, "Inscription", 150000, 150000, "2025-09-15", "2025-09-12", "payé", "REC-002"),
            (3, "Inscription", 150000, 80000, "2025-09-15", "2025-09-14", "partiel", "REC-003"),
            (4, "Inscription", 150000, 0, "2025-09-15", None, "impayé", None),
            (5, "Inscription", 150000, 75000, "2025-09-15", "2025-09-13", "partiel", "REC-004"),
        ]
        for fd in fee_data:
            c.execute("INSERT INTO fees (student_id,fee_type,amount,amount_paid,due_date,payment_date,status,receipt_number) VALUES (?,?,?,?,?,?,?,?)", fd)

    conn.commit()

    # ── Security: users table + default admin ──
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT DEFAULT 'viewer' CHECK(role IN ('admin','editor','viewer')),
        full_name TEXT,
        is_active INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now')),
        last_login TEXT
    )''')
    conn.commit()

    existing_admin = c.execute("SELECT id FROM users WHERE role='admin'").fetchone()
    if not existing_admin:
        c.execute("INSERT INTO users (username, password_hash, role, full_name) VALUES (?,?,'admin',?)",
                  ('admin', generate_password_hash('congoschool2025!', 'pbkdf2:sha256', 260000), 'Administrateur'))
        conn.commit()

    conn.close()


init_db()

# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────
def badge(status, cat="default"):
    colors = {
        "actif": "badge-success", "inactif": "badge-danger",
        "payé": "badge-success", "partiel": "badge-warning", "impayé": "badge-danger",
        "présent": "badge-success", "absent": "badge-danger", "retard": "badge-warning",
        "M": "badge-info", "F": "badge-pink",
    }
    cls = colors.get(status.lower() if status else "", "badge-secondary")
    return f'<span class="badge {cls}">{status or "-"}</span>'


def status_select(name, current="", options=None):
    if options is None:
        options = ["actif", "inactif"]
    parts = [f'<select name="{name}" class="form-select">']
    for o in options:
        sel = ' selected' if o == current else ''
        parts.append(f'<option value="{o}"{sel}>{o}</option>')
    parts.append('</select>')
    return '\n'.join(parts)


# ──────────────────────────────────────────────
# BASE TEMPLATE (sidebar + shell)
# ──────────────────────────────────────────────
BASE = '''<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CongoSchool - Gestion Scolaire</title>
<style>
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:'Segoe UI',Tahoma,Geneva,Verdana,sans-serif;background:#f0f2f5;color:#2d3436;min-height:100vh;}
.sidebar{width:250px;background:#1a1a2e;color:#fff;position:fixed;top:0;left:0;height:100vh;overflow-y:auto;z-index:1000;transition:transform .3s;}
.sidebar .logo{padding:20px;text-align:center;border-bottom:1px solid #16213e;font-size:22px;font-weight:700;}
.sidebar .logo span{color:#e94560;}
.sidebar nav a{display:flex;align-items:center;padding:12px 20px;color:#a0a0b8;text-decoration:none;transition:all .2s;font-size:14px;border-left:3px solid transparent;}
.sidebar nav a:hover,.sidebar nav a.active{background:#16213e;color:#fff;border-left-color:#e94560;}
.sidebar nav a .icon{width:24px;text-align:center;margin-right:12px;font-size:16px;}
.main{margin-left:250px;padding:24px;min-height:100vh;}
.topbar{display:flex;justify-content:space-between;align-items:center;margin-bottom:24px;}
.topbar h1{font-size:24px;color:#1a1a2e;}
.topbar .date{color:#636e72;font-size:14px;}
.toggle-btn{display:none;background:#e94560;color:#fff;border:none;padding:10px 14px;border-radius:6px;font-size:18px;cursor:pointer;position:fixed;top:12px;left:12px;z-index:1100;}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:16px;margin-bottom:24px;}
.card{background:#fff;border-radius:12px;padding:20px;box-shadow:0 2px 8px rgba(0,0,0,.06);}
.card .card-label{font-size:13px;color:#636e72;margin-bottom:6px;}
.card .card-value{font-size:28px;font-weight:700;}
.card .card-icon{font-size:28px;float:right;opacity:.3;}
.card-blue{border-left:4px solid #0984e3;} .card-blue .card-value{color:#0984e3;}
.card-green{border-left:4px solid #10b981;} .card-green .card-value{color:#10b981;}
.card-red{border-left:4px solid #e94560;} .card-red .card-value{color:#e94560;}
.card-yellow{border-left:4px solid #f59e0b;} .card-yellow .card-value{color:#f59e0b;}
.card-purple{border-left:4px solid #6c5ce7;} .card-purple .card-value{color:#6c5ce7;}
.toolbar{display:flex;gap:10px;margin-bottom:16px;flex-wrap:wrap;align-items:center;}
.btn{padding:8px 16px;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:600;text-decoration:none;display:inline-flex;align-items:center;gap:6px;transition:all .2s;}
.btn:hover{opacity:.85;transform:translateY(-1px);}
.btn-primary{background:#e94560;color:#fff;} .btn-success{background:#10b981;color:#fff;}
.btn-warning{background:#f59e0b;color:#fff;} .btn-danger{background:#ef4444;color:#fff;}
.btn-info{background:#0984e3;color:#fff;} .btn-secondary{background:#636e72;color:#fff;}
.btn-sm{padding:5px 10px;font-size:12px;}
.table-wrap{background:#fff;border-radius:12px;box-shadow:0 2px 8px rgba(0,0,0,.06);overflow-x:auto;}
table{width:100%;border-collapse:collapse;}
thead{background:#1a1a2e;color:#fff;}
th{padding:12px 14px;text-align:left;font-size:13px;font-weight:600;}
td{padding:10px 14px;border-bottom:1px solid #eee;font-size:13px;}
tbody tr:hover{background:#f8f9fa;}
.badge{padding:4px 10px;border-radius:20px;font-size:11px;font-weight:600;display:inline-block;}
.badge-success{background:#d1fae5;color:#065f46;} .badge-danger{background:#fee2e2;color:#991b1b;}
.badge-warning{background:#fef3c7;color:#92400e;} .badge-info{background:#dbeafe;color:#1e40af;}
.badge-secondary{background:#e5e7eb;color:#4b5563;} .badge-pink{background:#fce7f3;color:#9d174d;}
input,select,textarea{padding:8px 12px;border:1px solid #d1d5db;border-radius:6px;font-size:13px;font-family:inherit;width:100%;}
input:focus,select:focus,textarea:focus{outline:none;border-color:#e94560;box-shadow:0 0 0 3px rgba(233,69,96,.15);}
label{font-size:13px;font-weight:600;color:#374151;margin-bottom:4px;display:block;}
.form-group{margin-bottom:14px;}
.modal-overlay{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.5);z-index:2000;justify-content:center;align-items:center;}
.modal-overlay.show{display:flex;}
.modal{background:#fff;border-radius:12px;padding:28px;width:90%;max-width:540px;max-height:90vh;overflow-y:auto;position:relative;}
.modal h2{margin-bottom:20px;color:#1a1a2e;font-size:20px;}
.modal .close-btn{position:absolute;top:12px;right:16px;background:none;border:none;font-size:22px;cursor:pointer;color:#636e72;}
.modal .close-btn:hover{color:#e94560;}
.flash{padding:10px 16px;border-radius:8px;margin-bottom:16px;font-size:13px;}
.flash-success{background:#d1fae5;color:#065f46;} .flash-error{background:#fee2e2;color:#991b1b;}
.search-input{min-width:220px;}
@media(max-width:768px){
.sidebar{transform:translateX(-100%);} .sidebar.open{transform:translateX(0);}
.toggle-btn{display:block;} .main{margin-left:0;padding:16px;padding-top:56px;}
.cards{grid-template-columns:1fr 1fr;} .toolbar{flex-direction:column;align-items:stretch;}
}
</style>
</head>
<body>
<button class="toggle-btn" onclick="document.querySelector('.sidebar').classList.toggle('open')">☰</button>
<aside class="sidebar">
  <div class="logo"><span>⚔</span> CongoSchool</div>
  <nav>
    <a href="/" class="{{'active' if page=='dashboard'}}"><span class="icon">📊</span> Tableau de bord</a>
    <a href="/students" class="{{'active' if page=='students'}}"><span class="icon">🎓</span> Élèves</a>
    <a href="/teachers" class="{{'active' if page=='teachers'}}"><span class="icon">👨‍🏫</span> Enseignants</a>
    <a href="/classes" class="{{'active' if page=='classes'}}"><span class="icon">🏫</span> Classes</a>
    <a href="/subjects" class="{{'active' if page=='subjects'}}"><span class="icon">📚</span> Matières</a>
    <a href="/grades" class="{{'active' if page=='grades'}}"><span class="icon">📝</span> Notes</a>
    <a href="/attendance" class="{{'active' if page=='attendance'}}"><span class="icon">✅</span> Présences</a>
    <a href="/timetable" class="{{'active' if page=='timetable'}}"><span class="icon">📅</span> Emploi du temps</a>
    <a href="/fees" class="{{'active' if page=='fees'}}"><span class="icon">💰</span> Frais scolaires</a>
    {% if session.get('user_role') == 'admin' %}<a href="/admin/users" class="{{'active' if page=='admin_users'}}"><span class="icon">👤</span> Utilisateurs</a>{% endif %}
    <a href="/logout" style="margin-top:20px;border-top:1px solid #16213e;padding-top:12px;color:#e94560;"><span class="icon">🚪</span> Déconnexion</a>
  </nav>
</aside>
<div class="main">
  <div class="topbar">
    <h1>{{title}}</h1>
    <div style="display:flex;align-items:center;gap:12px;">
      <span style="font-size:12px;color:#636e72;">👤 {{session.get('user_name','?')}} ({{session.get('user_role','?')}})</span>
      <form method="post" action="/set_cycle" style="display:flex;gap:4px;background:#16213e;border-radius:8px;padding:3px;">
        <input type="hidden" name="cycle" value="primaire" id="cycle-hidden">
        <button type="button" onclick="document.getElementById('cycle-hidden').value='primaire';this.form.submit()" style="padding:8px 16px;border:none;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;{% if cycle == 'primaire' %}background:#e94560;color:#fff;{% else %}background:transparent;color:#a0a0b8;{% endif %}">🏫 Primaire</button>
        <button type="button" onclick="document.getElementById('cycle-hidden').value='secondaire';this.form.submit()" style="padding:8px 16px;border:none;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;{% if cycle == 'secondaire' %}background:#e94560;color:#fff;{% else %}background:transparent;color:#a0a0b8;{% endif %}">📖 Secondaire</button>
      </form>
      <span class="date">{{today_str}}</span>
    </div>
  </div>
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}
      {% for cat, msg in messages %}<div class="flash flash-{{cat}}">{{msg}}</div>{% endfor %}
    {% endif %}
  {% endwith %}
  {% block content %}{% endblock %}
</div>
{% block modals %}{% endblock %}
<script>
function openModal(id){document.getElementById(id).classList.add('show');}
function closeModal(id){document.getElementById(id).classList.remove('show');}
</script>
</body>
</html>'''


def render(page, title, content, modals=""):
    today_str = date.today().strftime("%A %d %B %Y")
    cycle = session.get('cycle', 'primaire')
    tpl = BASE.replace("{% block content %}{% endblock %}", "{% block content %}" + content + "{% endblock %}")
    tpl = tpl.replace("{% block modals %}{% endblock %}", "{% block modals %}" + modals + "{% endblock %}")
    return render_template_string(tpl, page=page, title=title, today_str=today_str, cycle=cycle)


# ──────────────────────────────────────────────
# AUTHENTICATION & SECURITY
# ──────────────────────────────────────────────
def current_user():
    if 'user_id' not in session:
        return None
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id=? AND is_active=1", (session['user_id'],)).fetchone()
    conn.close()
    return user

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = current_user()
        if not user or user['role'] != 'admin':
            flash("Accès réservé à l'administrateur.", "error")
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

@app.before_request
def auth_guard():
    # Exempt login, static, and favicon
    if request.path in ('/login', '/favicon.ico') or request.path.startswith('/static'):
        return None
    if 'user_id' not in session:
        return redirect(url_for('login'))
    # Check user still exists and is active
    conn = get_db()
    user = conn.execute("SELECT id FROM users WHERE id=? AND is_active=1", (session['user_id'],)).fetchone()
    conn.close()
    if not user:
        session.clear()
        return redirect(url_for('login'))
    # Write protection: POST/DELETE requires at least 'editor' role
    if request.method in ('POST', 'DELETE') and request.path != '/login':
        role = session.get('user_role', 'viewer')
        if role == 'viewer' and not request.path.startswith('/admin'):
            flash("Accès en lecture seule. Contactez l'administrateur pour modifier.", "error")
            return redirect(url_for('dashboard'))
    return None

LOGIN_STYLE = '''
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{min-height:100vh;display:flex;align-items:center;justify-content:center;background:linear-gradient(135deg,#0f0c29,#302b63,#24243e);font-family:Segoe UI,sans-serif}
.login-box{background:#fff;border-radius:16px;padding:40px;width:90%;max-width:420px;box-shadow:0 20px 60px rgba(0,0,0,.4)}
.login-box .logo{text-align:center;margin-bottom:8px;font-size:28px;font-weight:800;color:#1a1a2e}
.login-box .logo span{color:#e94560}
.login-box .subtitle{text-align:center;color:#636e72;font-size:13px;margin-bottom:28px}
.login-box label{font-size:13px;font-weight:600;color:#374151;margin-bottom:6px;display:block}
.login-box input{width:100%;padding:12px 14px;border:2px solid #d1d5db;border-radius:10px;font-size:14px;transition:border .2s;margin-bottom:16px}
.login-box input:focus{outline:none;border-color:#e94560;box-shadow:0 0 0 3px rgba(233,69,96,.15)}
.login-box button{width:100%;padding:14px;border:none;border-radius:10px;background:linear-gradient(135deg,#e94560,#c0392b);color:#fff;font-size:15px;font-weight:700;cursor:pointer;transition:transform .2s}
.login-box button:hover{transform:translateY(-2px);box-shadow:0 8px 20px rgba(233,69,96,.3)}
.login-box .error{text-align:center;color:#e94560;font-size:13px;margin-bottom:12px;padding:8px;background:#fee2e2;border-radius:8px}
.login-box .footer{text-align:center;margin-top:20px;font-size:11px;color:#aaa}
</style>
'''

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        pwd = request.form.get('password', '')
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE username=? AND is_active=1", (username,)).fetchone()
        conn.close()
        if user and check_password_hash(user['password_hash'], pwd):
            session['user_id'] = user['id']
            session['user_role'] = user['role']
            session['user_name'] = user['full_name'] or user['username']
            # Update last login
            conn = get_db()
            conn.execute("UPDATE users SET last_login=datetime('now') WHERE id=?", (user['id'],))
            conn.commit()
            conn.close()
            return redirect(url_for('dashboard'))
        else:
            error = "Nom d'utilisateur ou mot de passe incorrect."
    err_html = f'<div class="error">{error}</div>' if error else ''
    return render_template_string(LOGIN_STYLE + '''
<body>
<div class="login-box">
  <div class="logo"><span>⚔</span> CongoSchool</div>
  <div class="subtitle">Connectez-vous pour accéder au système</div>
  ''' + err_html + '''
  <form method="post">
    <label>Nom d'utilisateur</label>
    <input type="text" name="username" placeholder="Entrez votre identifiant" required autofocus>
    <label>Mot de passe</label>
    <input type="password" name="password" placeholder="Entrez votre mot de passe" required>
    <button type="submit">🔐 Se connecter</button>
  </form>
  <div class="footer">CongoSchool &copy; 2025 — Accès sécurisé</div>
</div>
</body>
''')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/admin/users')
@admin_required
def admin_users():
    conn = get_db()
    users = conn.execute("SELECT id, username, full_name, role, is_active, created_at, last_login FROM users ORDER BY id").fetchall()
    conn.close()
    rows = ""
    for u in users:
        role_badge = "badge-danger" if u['role']=='admin' else ("badge-info" if u['role']=='editor' else "badge-secondary")
        status_badge = "badge-success" if u['is_active'] else "badge-danger"
        status_text = "Actif" if u['is_active'] else "Désactivé"
        rows += f'''<tr>
            <td>{u['id']}</td>
            <td><b>{u['username']}</b></td>
            <td>{u['full_name'] or '-'}</td>
            <td><span class="badge {role_badge}">{u['role'].upper()}</span></td>
            <td><span class="badge {status_badge}">{status_text}</span></td>
            <td>{u['last_login'] or '-'}</td>
            <td>
                <form method="post" action="/admin/users/toggle" style="display:inline">
                    <input type="hidden" name="uid" value="{u['id']}">
                    <button type="submit" class="btn btn-sm {'btn-danger' if u['is_active'] else 'btn-success'}">{'Désactiver' if u['is_active'] else 'Activer'}</button>
                </form>
                <form method="post" action="/admin/users/reset_pwd" style="display:inline" onsubmit="return confirm('Réinitialiser le mot de passe ?')">
                    <input type="hidden" name="uid" value="{u['id']}">
                    <button type="submit" class="btn btn-sm btn-warning">🔑 Reset</button>
                </form>
                <form method="post" action="/admin/users/delete" style="display:inline" onsubmit="return confirm('Supprimer cet utilisateur ?')">
                    <input type="hidden" name="uid" value="{u['id']}">
                    <button type="submit" class="btn btn-sm btn-danger">🗑️</button>
                </form>
            </td>
        </tr>'''
    content = f'''
    <div class="toolbar">
        <h2>👤 Gestion des utilisateurs</h2>
        <button class="btn btn-primary" onclick="openModal('addUserModal')">+ Nouvel utilisateur</button>
    </div>
    <div class="table-wrap"><table>
        <thead><tr><th>ID</th><th>Utilisateur</th><th>Nom complet</th><th>Rôle</th><th>Statut</th><th>Dernière connexion</th><th>Actions</th></tr></thead>
        <tbody>{rows}</tbody>
    </table></div>
    <div id="addUserModal" class="modal-overlay">
        <div class="modal">
            <h2>➕ Nouvel utilisateur</h2>
            <form method="post" action="/admin/users/add">
                <div class="form-group"><label>Utilisateur</label><input type="text" name="username" required></div>
                <div class="form-group"><label>Nom complet</label><input type="text" name="full_name"></div>
                <div class="form-group"><label>Mot de passe</label><input type="password" name="password" required></div>
                <div class="form-group"><label>Rôle</label><select name="role"><option value="viewer">Viewer (lecture seule)</option><option value="editor">Éditeur (créer/modifier)</option><option value="admin">Admin (tout)</option></select></div>
                <div style="display:flex;gap:10px;justify-content:flex-end;margin-top:20px">
                    <button type="button" class="btn btn-secondary" onclick="closeModal('addUserModal')">Annuler</button>
                    <button type="submit" class="btn btn-primary">Créer</button>
                </div>
            </form>
            <button class="close-btn" onclick="closeModal('addUserModal')">&times;</button>
        </div>
    </div>'''
    return render("admin_users", "Gestion des utilisateurs", content)

@app.route('/admin/users/add', methods=['POST'])
@admin_required
def admin_add_user():
    username = request.form.get('username', '').strip()
    full_name = request.form.get('full_name', '').strip()
    password = request.form.get('password', '')
    role = request.form.get('role', 'viewer')
    if not username or not password:
        flash("Nom d'utilisateur et mot de passe requis.", "error")
        return redirect(url_for('admin_users'))
    if len(password) < 6:
        flash("Le mot de passe doit avoir au moins 6 caractères.", "error")
        return redirect(url_for('admin_users'))
    conn = get_db()
    try:
        conn.execute("INSERT INTO users (username, password_hash, role, full_name) VALUES (?,?,?,?)",
                     (username, generate_password_hash(password, 'pbkdf2:sha256', 260000), role, full_name or None))
        conn.commit()
        flash(f"Utilisateur '{username}' créé avec succès.", "success")
    except Exception:
        flash("Cet utilisateur existe déjà.", "error")
    conn.close()
    return redirect(url_for('admin_users'))

@app.route('/admin/users/toggle', methods=['POST'])
@admin_required
def admin_toggle_user():
    uid = request.form.get('uid', type=int)
    conn = get_db()
    if uid == session.get('user_id'):
        flash("Vous ne pouvez pas vous désactiver vous-même.", "error")
    else:
        conn.execute("UPDATE users SET is_active = CASE WHEN is_active=1 THEN 0 ELSE 1 END WHERE id=?", (uid,))
        conn.commit()
        flash("Statut mis à jour.", "success")
    conn.close()
    return redirect(url_for('admin_users'))

@app.route('/admin/users/reset_pwd', methods=['POST'])
@admin_required
def admin_reset_pwd():
    uid = request.form.get('uid', type=int)
    new_pwd = 'congoschool2025!'
    conn = get_db()
    conn.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(new_pwd, 'pbkdf2:sha256', 260000), uid))
    conn.commit()
    conn.close()
    flash(f"Mot de passe réinitialisé pour l'utilisateur {uid} : {new_pwd}", "success")
    return redirect(url_for('admin_users'))

@app.route('/admin/users/delete', methods=['POST'])
@admin_required
def admin_delete_user():
    uid = request.form.get('uid', type=int)
    if uid == session.get('user_id'):
        flash("Vous ne pouvez pas supprimer votre propre compte.", "error")
    else:
        conn = get_db()
        conn.execute("DELETE FROM users WHERE id=?", (uid,))
        conn.commit()
        conn.close()
        flash("Utilisateur supprimé.", "success")
    return redirect(url_for('admin_users'))


# ──────────────────────────────────────────────
# ROUTES
# ──────────────────────────────────────────────

@app.route('/set_cycle', methods=['POST'])
def set_cycle():
    cycle = request.form.get('cycle', 'primaire')
    if cycle in ('primaire', 'secondaire'):
        session['cycle'] = cycle
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/')
def dashboard():
    conn = get_db()
    cycle = session.get('cycle', 'primaire')
    nb_students = conn.execute("SELECT COUNT(*) FROM students WHERE cycle=?", (cycle,)).fetchone()[0]
    nb_teachers = conn.execute("SELECT COUNT(*) FROM teachers").fetchone()[0]
    nb_classes = conn.execute("SELECT COUNT(*) FROM classes WHERE cycle=?", (cycle,)).fetchone()[0]
    nb_subjects = conn.execute("SELECT COUNT(*) FROM subjects").fetchone()[0]

    rows = conn.execute("""
        SELECT c.name, COUNT(s.id) as nb FROM classes c
        LEFT JOIN students s ON s.class_id = c.id AND s.cycle=c.cycle
        WHERE c.cycle=? GROUP BY c.id ORDER BY c.name
    """, (cycle,)).fetchall()

    recent = conn.execute("""
        SELECT s.*, c.name as class_name FROM students s
        LEFT JOIN classes c ON c.id = s.class_id
        WHERE s.cycle=? ORDER BY s.id DESC LIMIT 5
    """, (cycle,)).fetchall()
    conn.close()

    class_rows = "".join(
        f"<tr><td>{r['name']}</td><td>{r['nb']} élèves</td></tr>" for r in rows
    )
    recent_rows = "".join(
        f"<tr><td>{r['matricule']}</td><td>{r['first_name']} {r['last_name']}</td>"
        f"<td>{r['class_name'] or '-'}</td><td>{badge(r['status'])}</td></tr>"
        for r in recent
    )

    content = f'''
    <div class="cards">
      <div class="card card-blue"><span class="card-icon">🎓</span><div class="card-label">Élèves</div><div class="card-value">{nb_students}</div></div>
      <div class="card card-green"><span class="card-icon">👨‍🏫</span><div class="card-label">Enseignants</div><div class="card-value">{nb_teachers}</div></div>
      <div class="card card-purple"><span class="card-icon">🏫</span><div class="card-label">Classes</div><div class="card-value">{nb_classes}</div></div>
      <div class="card card-yellow"><span class="card-icon">📚</span><div class="card-label">Matières</div><div class="card-value">{nb_subjects}</div></div>
    </div>
    <div style="margin:16px 0;padding:16px 20px;background:linear-gradient(135deg,#1a1a2e,#16213e);border-radius:12px;display:flex;align-items:center;justify-content:space-between;">
      <div style="color:#fff"><div style="font-size:16px;font-weight:700">📦 Clôturer l'année scolaire</div><div style="font-size:12px;opacity:.7;margin-top:4px">Exporte toutes les données en ZIP, puis vide les données pour la nouvelle année. <b style="color:#ff6b6b">Irreversible !</b></div></div>
      <form id="closeYearForm" onsubmit="return confirm(\'⚠️ Êtes-vous sûr ?\\n\\nToutes les données (élèves, notes, présences, frais) seront archivées et supprimées.\\nLes enseignants, classes et matières seront conservés.\\n\\nUn fichier ZIP de sauvegarde sera téléchargé.\')" method="post" action="/close_year">
        <input type="text" name="confirm_text" placeholder="Tapez CONGO pour confirmer" required
          style="padding:8px 12px;border-radius:8px;border:2px solid #e94560;background:#0f3460;color:#fff;font-size:13px;width:200px;margin-right:10px" />
        <button type="submit" style="padding:10px 24px;border:none;border-radius:8px;background:#e94560;color:#fff;font-size:14px;font-weight:700;cursor:pointer">Clôturer & Exporter</button>
      </form>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;">
      <div class="table-wrap"><table><thead><tr><th>Classe</th><th>Effectif</th></tr></thead><tbody>{class_rows}</tbody></table></div>
      <div class="table-wrap"><table><thead><tr><th>Matricule</th><th>Nom</th><th>Classe</th><th>Statut</th></tr></thead><tbody>{recent_rows}</tbody></table></div>
    </div>'''
    return render("dashboard", "Tableau de bord", content)


# ────────────── STUDENTS ──────────────
@app.route('/students')
def students_list():
    conn = get_db()
    cycle = session.get('cycle', 'primaire')
    q = request.args.get('q', '').strip()
    cid = request.args.get('class_id', '')
    sql = "SELECT s.*, c.name as class_name FROM students s LEFT JOIN classes c ON c.id = s.class_id WHERE s.cycle=?"
    params = [cycle]
    if q:
        sql += " AND (s.first_name LIKE ? OR s.last_name LIKE ? OR s.matricule LIKE ?)"
        params += [f"%{q}%"] * 3
    if cid:
        sql += " AND s.class_id = ?"
        params.append(int(cid))
    sql += " ORDER BY s.last_name, s.first_name"
    rows = conn.execute(sql, params).fetchall()
    classes = conn.execute("SELECT id, name FROM classes WHERE cycle=? ORDER BY name", (cycle,)).fetchall()
    conn.close()

    class_opts = '<option value="">Toutes</option>' + "".join(
        f'<option value="{c["id"]}" {"selected" if str(c["id"])==cid else ""}>{c["name"]}</option>' for c in classes
    )
    rows_html = ""
    for r in rows:
        rows_html += f'''<tr>
            <td>{r['matricule']}</td><td>{r['first_name']} {r['last_name']}</td>
            <td>{badge(r['gender'])}</td><td>{r['class_name'] or '-'}</td>
            <td>{r['parent_name'] or '-'}</td><td>{badge(r['status'])}</td>
            <td>
              <a href="/students/edit/{r['id']}" class="btn btn-warning btn-sm">✏️</a>
              <a href="/students/delete/{r['id']}" class="btn btn-danger btn-sm" onclick="return confirm('Supprimer?')">🗑️</a>
            </td></tr>'''

    content = f'''
    <div class="toolbar">
      <form method="get" style="display:flex;gap:8px;flex-wrap:wrap;">
        <input type="text" name="q" value="{q}" placeholder="Rechercher..." class="search-input">
        <select name="class_id" onchange="this.form.submit()">{class_opts}</select>
      </form>
      <a href="/export/students" class="btn btn-success btn-sm">📊 Excel</a>
      <button class="btn btn-primary" onclick="openModal('modal-student')">+ Nouvel élève</button>
    </div>
    <div class="table-wrap"><table>
      <thead><tr><th>Matricule</th><th>Nom complet</th><th>Sexe</th><th>Classe</th><th>Parent</th><th>Statut</th><th>Actions</th></tr></thead>
      <tbody>{rows_html}</tbody>
    </table></div>'''

    conn2 = get_db()
    classes2 = conn2.execute("SELECT id, name FROM classes ORDER BY name").fetchall()
    conn2.close()
    class_opts2 = '<option value="">-- Choisir --</option>' + "".join(
        f'<option value="{c["id"]}">{c["name"]}</option>' for c in classes2
    )

    modal = f'''
    <div class="modal-overlay" id="modal-student">
      <div class="modal">
        <button class="close-btn" onclick="closeModal('modal-student')">&times;</button>
        <h2>Nouvel élève</h2>
        <form method="post" action="/students/add">
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
            <div class="form-group"><label>Matricule</label><input name="matricule" required></div>
            <div class="form-group"><label>Sexe</label><select name="gender" class="form-select"><option value="M">Masculin</option><option value="F">Féminin</option></select></div>
            <div class="form-group"><label>Prénom</label><input name="first_name" required></div>
            <div class="form-group"><label>Nom</label><input name="last_name" required></div>
            <div class="form-group"><label>Date de naissance</label><input type="date" name="date_of_birth"></div>
            <div class="form-group"><label>Classe</label><select name="class_id" class="form-select">{class_opts2}</select></div>
            <div class="form-group"><label>Nom du parent</label><input name="parent_name"></div>
            <div class="form-group"><label>Téléphone parent</label><input name="parent_phone"></div>
            <div class="form-group"><label>Adresse</label><input name="address"></div>
            <div class="form-group"><label>Email</label><input type="email" name="email"></div>
          </div>
          <div style="margin-top:14px;display:flex;gap:8px;justify-content:flex-end;">
            <button type="button" class="btn btn-secondary" onclick="closeModal('modal-student')">Annuler</button>
            <button type="submit" class="btn btn-primary">Enregistrer</button>
          </div>
        </form>
      </div>
    </div>'''
    return render("students", "Élèves", content, modal)


@app.route('/students/add', methods=['POST'])
def students_add():
    try:
        conn = get_db()
        cycle = session.get('cycle', 'primaire')
        conn.execute("""INSERT INTO students (matricule,first_name,last_name,date_of_birth,gender,class_id,
            parent_name,parent_phone,address,email,enrollment_date,status,cycle)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (request.form.get('matricule'), request.form.get('first_name'), request.form.get('last_name'),
             request.form.get('date_of_birth') or None, request.form.get('gender'),
             int(request.form.get('class_id') or 0) or None,
             request.form.get('parent_name'), request.form.get('parent_phone'),
             request.form.get('address'), request.form.get('email'),
             date.today().isoformat(), 'actif', cycle))
        conn.commit()
        conn.close()
        flash("Élève ajouté avec succès", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('students_list'))


@app.route('/students/edit/<int:sid>', methods=['GET', 'POST'])
def students_edit(sid):
    conn = get_db()
    s = conn.execute("SELECT * FROM students WHERE id=?", (sid,)).fetchone()
    if request.method == 'POST':
        try:
            conn.execute("""UPDATE students SET matricule=?,first_name=?,last_name=?,date_of_birth=?,gender=?,
                class_id=?,parent_name=?,parent_phone=?,address=?,email=?,status=? WHERE id=?""",
                (request.form.get('matricule'), request.form.get('first_name'), request.form.get('last_name'),
                 request.form.get('date_of_birth') or None, request.form.get('gender'),
                 int(request.form.get('class_id') or 0) or None,
                 request.form.get('parent_name'), request.form.get('parent_phone'),
                 request.form.get('address'), request.form.get('email'),
                 request.form.get('status'), sid))
            conn.commit()
            conn.close()
            flash("Élève modifié", "success")
            return redirect(url_for('students_list'))
        except Exception as e:
            flash(f"Erreur: {e}", "error")
    classes = conn.execute("SELECT id,name FROM classes ORDER BY name").fetchall()
    conn.close()
    class_opts = '<option value="">-- Aucune --</option>' + "".join(
        f'<option value="{c["id"]}" {"selected" if c["id"]==s["class_id"] else ""}>{c["name"]}</option>' for c in classes
    )
    content = f'''
    <div class="card" style="max-width:600px;">
      <h2 style="margin-bottom:16px;">Modifier l'élève</h2>
      <form method="post">
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
          <div class="form-group"><label>Matricule</label><input name="matricule" value="{s['matricule']}" required></div>
          <div class="form-group"><label>Sexe</label><select name="gender" class="form-select"><option value="M" {"selected" if s['gender']=='M' else ""}>Masculin</option><option value="F" {"selected" if s['gender']=='F' else ""}>Féminin</option></select></div>
          <div class="form-group"><label>Prénom</label><input name="first_name" value="{s['first_name']}" required></div>
          <div class="form-group"><label>Nom</label><input name="last_name" value="{s['last_name']}" required></div>
          <div class="form-group"><label>Date de naissance</label><input type="date" name="date_of_birth" value="{s['date_of_birth'] or ''}"></div>
          <div class="form-group"><label>Classe</label><select name="class_id" class="form-select">{class_opts}</select></div>
          <div class="form-group"><label>Nom du parent</label><input name="parent_name" value="{s['parent_name'] or ''}"></div>
          <div class="form-group"><label>Téléphone parent</label><input name="parent_phone" value="{s['parent_phone'] or ''}"></div>
          <div class="form-group"><label>Adresse</label><input name="address" value="{s['address'] or ''}"></div>
          <div class="form-group"><label>Email</label><input type="email" name="email" value="{s['email'] or ''}"></div>
          <div class="form-group"><label>Statut</label>{status_select('status', s['status'] or 'actif')}</div>
        </div>
        <div style="margin-top:14px;display:flex;gap:8px;">
          <button type="submit" class="btn btn-success">Sauvegarder</button>
          <a href="/students" class="btn btn-secondary">Annuler</a>
        </div>
      </form>
    </div>'''
    return render("students", "Modifier élève", content)


@app.route('/students/delete/<int:sid>')
def students_delete(sid):
    try:
        conn = get_db()
        conn.execute("DELETE FROM students WHERE id=?", (sid,))
        conn.commit()
        conn.close()
        flash("Élève supprimé", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('students_list'))


# ────────────── TEACHERS ──────────────
@app.route('/teachers')
def teachers_list():
    conn = get_db()
    rows = conn.execute("SELECT * FROM teachers ORDER BY last_name, first_name").fetchall()
    conn.close()

    rows_html = ""
    for r in rows:
        rows_html += f'''<tr>
            <td>{r['first_name']} {r['last_name']}</td>
            <td>{r['subject'] or '-'}</td><td>{r['phone'] or '-'}</td><td>{r['email'] or '-'}</td>
            <td>{badge(r['status'])}</td>
            <td>
              <a href="/teachers/edit/{r['id']}" class="btn btn-warning btn-sm">✏️</a>
              <a href="/teachers/delete/{r['id']}" class="btn btn-danger btn-sm" onclick="return confirm('Supprimer?')">🗑️</a>
            </td></tr>'''

    content = f'''
    <div class="toolbar">
      <div></div>
      <a href="/export/teachers" class="btn btn-success btn-sm">📊 Excel</a>
      <button class="btn btn-primary" onclick="openModal('modal-teacher')">+ Nouvel enseignant</button>
    </div>
    <div class="table-wrap"><table>
      <thead><tr><th>Nom</th><th>Matière</th><th>Téléphone</th><th>Email</th><th>Statut</th><th>Actions</th></tr></thead>
      <tbody>{rows_html}</tbody>
    </table></div>'''

    modal = '''
    <div class="modal-overlay" id="modal-teacher">
      <div class="modal">
        <button class="close-btn" onclick="closeModal('modal-teacher')">&times;</button>
        <h2>Nouvel enseignant</h2>
        <form method="post" action="/teachers/add">
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
            <div class="form-group"><label>Prénom</label><input name="first_name" required></div>
            <div class="form-group"><label>Nom</label><input name="last_name" required></div>
            <div class="form-group"><label>Téléphone</label><input name="phone"></div>
            <div class="form-group"><label>Email</label><input type="email" name="email"></div>
            <div class="form-group"><label>Matière</label><input name="subject"></div>
            <div class="form-group"><label>Adresse</label><input name="address"></div>
            <div class="form-group"><label>Date d'embauche</label><input type="date" name="hire_date"></div>
            <div class="form-group"><label>Salaire (CDF)</label><input type="number" name="salary" value="0"></div>
          </div>
          <div style="margin-top:14px;display:flex;gap:8px;justify-content:flex-end;">
            <button type="button" class="btn btn-secondary" onclick="closeModal('modal-teacher')">Annuler</button>
            <button type="submit" class="btn btn-primary">Enregistrer</button>
          </div>
        </form>
      </div>
    </div>'''
    return render("teachers", "Enseignants", content, modal)


@app.route('/teachers/add', methods=['POST'])
def teachers_add():
    try:
        conn = get_db()
        conn.execute("""INSERT INTO teachers (first_name,last_name,phone,email,subject,address,hire_date,status,salary)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            (request.form.get('first_name'), request.form.get('last_name'),
             request.form.get('phone'), request.form.get('email'), request.form.get('subject'),
             request.form.get('address'), request.form.get('hire_date') or None,
             'actif', float(request.form.get('salary') or 0)))
        conn.commit()
        conn.close()
        flash("Enseignant ajouté", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('teachers_list'))


@app.route('/teachers/edit/<int:tid>', methods=['GET', 'POST'])
def teachers_edit(tid):
    conn = get_db()
    t = conn.execute("SELECT * FROM teachers WHERE id=?", (tid,)).fetchone()
    if request.method == 'POST':
        try:
            conn.execute("""UPDATE teachers SET first_name=?,last_name=?,phone=?,email=?,subject=?,
                address=?,hire_date=?,status=?,salary=? WHERE id=?""",
                (request.form.get('first_name'), request.form.get('last_name'),
                 request.form.get('phone'), request.form.get('email'), request.form.get('subject'),
                 request.form.get('address'), request.form.get('hire_date') or None,
                 request.form.get('status'), float(request.form.get('salary') or 0), tid))
            conn.commit()
            conn.close()
            flash("Enseignant modifié", "success")
            return redirect(url_for('teachers_list'))
        except Exception as e:
            flash(f"Erreur: {e}", "error")
    conn.close()
    content = f'''
    <div class="card" style="max-width:600px;">
      <h2 style="margin-bottom:16px;">Modifier l'enseignant</h2>
      <form method="post">
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
          <div class="form-group"><label>Prénom</label><input name="first_name" value="{t['first_name']}" required></div>
          <div class="form-group"><label>Nom</label><input name="last_name" value="{t['last_name']}" required></div>
          <div class="form-group"><label>Téléphone</label><input name="phone" value="{t['phone'] or ''}"></div>
          <div class="form-group"><label>Email</label><input type="email" name="email" value="{t['email'] or ''}"></div>
          <div class="form-group"><label>Matière</label><input name="subject" value="{t['subject'] or ''}"></div>
          <div class="form-group"><label>Adresse</label><input name="address" value="{t['address'] or ''}"></div>
          <div class="form-group"><label>Date d'embauche</label><input type="date" name="hire_date" value="{t['hire_date'] or ''}"></div>
          <div class="form-group"><label>Salaire</label><input type="number" name="salary" value="{t['salary']}"></div>
          <div class="form-group"><label>Statut</label>{status_select('status', t['status'] or 'actif')}</div>
        </div>
        <div style="margin-top:14px;display:flex;gap:8px;">
          <button type="submit" class="btn btn-success">Sauvegarder</button>
          <a href="/teachers" class="btn btn-secondary">Annuler</a>
        </div>
      </form>
    </div>'''
    return render("teachers", "Modifier enseignant", content)


@app.route('/teachers/delete/<int:tid>')
def teachers_delete(tid):
    try:
        conn = get_db()
        conn.execute("DELETE FROM teachers WHERE id=?", (tid,))
        conn.commit()
        conn.close()
        flash("Enseignant supprimé", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('teachers_list'))


# ────────────── CLASSES ──────────────
@app.route('/classes')
def classes_list():
    conn = get_db()
    cycle = session.get('cycle', 'primaire')
    rows = conn.execute("""
        SELECT c.*, t.first_name||' '||t.last_name as teacher_name, COUNT(s.id) as nb_students
        FROM classes c
        LEFT JOIN teachers t ON t.id = c.teacher_id
        LEFT JOIN students s ON s.class_id = c.id AND s.cycle=c.cycle
        WHERE c.cycle=?
        GROUP BY c.id ORDER BY c.name
    """, (cycle,)).fetchall()
    conn.close()

    rows_html = ""
    for r in rows:
        rows_html += f'''<tr>
            <td>{r['name']}</td><td>{r['level'] or '-'}</td>
            <td>{r['nb_students']}/{r['capacity']}</td>
            <td>{r['teacher_name'] or '-'}</td><td>{r['description'] or '-'}</td>
            <td>
              <a href="/classes/edit/{r['id']}" class="btn btn-warning btn-sm">✏️</a>
              <a href="/classes/delete/{r['id']}" class="btn btn-danger btn-sm" onclick="return confirm('Supprimer?')">🗑️</a>
            </td></tr>'''

    content = f'''
    <div class="toolbar">
      <div></div>
      <a href="/export/classes" class="btn btn-success btn-sm">📊 Excel</a>
      <button class="btn btn-primary" onclick="openModal('modal-class')">+ Nouvelle classe</button>
    </div>
    <div class="table-wrap"><table>
      <thead><tr><th>Nom</th><th>Niveau</th><th>Effectif/Capacité</th><th>Professeur principal</th><th>Description</th><th>Actions</th></tr></thead>
      <tbody>{rows_html}</tbody>
    </table></div>'''

    conn2 = get_db()
    teachers = conn2.execute("SELECT id, first_name, last_name FROM teachers ORDER BY last_name").fetchall()
    conn2.close()
    t_opts = '<option value="">-- Aucun --</option>' + "".join(
        f'<option value="{t["id"]}">{t["first_name"]} {t["last_name"]}</option>' for t in teachers
    )

    modal = f'''
    <div class="modal-overlay" id="modal-class">
      <div class="modal">
        <button class="close-btn" onclick="closeModal('modal-class')">&times;</button>
        <h2>Nouvelle classe</h2>
        <form method="post" action="/classes/add">
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
            <div class="form-group"><label>Nom</label><input name="name" required></div>
            <div class="form-group"><label>Niveau</label><input name="level"></div>
            <div class="form-group"><label>Capacité</label><input type="number" name="capacity" value="40"></div>
            <div class="form-group"><label>Professeur principal</label><select name="teacher_id" class="form-select">{t_opts}</select></div>
          </div>
          <div class="form-group"><label>Description</label><textarea name="description" rows="2"></textarea></div>
          <div style="margin-top:14px;display:flex;gap:8px;justify-content:flex-end;">
            <button type="button" class="btn btn-secondary" onclick="closeModal('modal-class')">Annuler</button>
            <button type="submit" class="btn btn-primary">Enregistrer</button>
          </div>
        </form>
      </div>
    </div>'''
    return render("classes", "Classes", content, modal)


@app.route('/classes/add', methods=['POST'])
def classes_add():
    try:
        conn = get_db()
        cycle = session.get('cycle', 'primaire')
        conn.execute("INSERT INTO classes (name,level,capacity,teacher_id,description,cycle) VALUES (?,?,?,?,?,?)",
            (request.form.get('name'), request.form.get('level'),
             int(request.form.get('capacity') or 40),
             int(request.form.get('teacher_id') or 0) or None,
             request.form.get('description'), cycle))
        conn.commit()
        conn.close()
        flash("Classe ajoutée", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('classes_list'))


@app.route('/classes/edit/<int:cid>', methods=['GET', 'POST'])
def classes_edit(cid):
    conn = get_db()
    c = conn.execute("SELECT * FROM classes WHERE id=?", (cid,)).fetchone()
    if request.method == 'POST':
        try:
            conn.execute("UPDATE classes SET name=?,level=?,capacity=?,teacher_id=?,description=? WHERE id=?",
                (request.form.get('name'), request.form.get('level'),
                 int(request.form.get('capacity') or 40),
                 int(request.form.get('teacher_id') or 0) or None,
                 request.form.get('description'), cid))
            conn.commit()
            conn.close()
            flash("Classe modifiée", "success")
            return redirect(url_for('classes_list'))
        except Exception as e:
            flash(f"Erreur: {e}", "error")
    teachers = conn.execute("SELECT id, first_name, last_name FROM teachers ORDER BY last_name").fetchall()
    conn.close()
    t_opts = '<option value="">-- Aucun --</option>' + "".join(
        f'<option value="{t["id"]}" {"selected" if t["id"]==c["teacher_id"] else ""}>{t["first_name"]} {t["last_name"]}</option>' for t in teachers
    )
    content = f'''
    <div class="card" style="max-width:600px;">
      <h2 style="margin-bottom:16px;">Modifier la classe</h2>
      <form method="post">
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
          <div class="form-group"><label>Nom</label><input name="name" value="{c['name']}" required></div>
          <div class="form-group"><label>Niveau</label><input name="level" value="{c['level'] or ''}"></div>
          <div class="form-group"><label>Capacité</label><input type="number" name="capacity" value="{c['capacity']}"></div>
          <div class="form-group"><label>Professeur principal</label><select name="teacher_id" class="form-select">{t_opts}</select></div>
        </div>
        <div class="form-group"><label>Description</label><textarea name="description" rows="2">{c['description'] or ''}</textarea></div>
        <div style="margin-top:14px;display:flex;gap:8px;">
          <button type="submit" class="btn btn-success">Sauvegarder</button>
          <a href="/classes" class="btn btn-secondary">Annuler</a>
        </div>
      </form>
    </div>'''
    return render("classes", "Modifier classe", content)


@app.route('/classes/delete/<int:cid>')
def classes_delete(cid):
    try:
        conn = get_db()
        conn.execute("DELETE FROM classes WHERE id=?", (cid,))
        conn.commit()
        conn.close()
        flash("Classe supprimée", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('classes_list'))


# ────────────── SUBJECTS ──────────────
@app.route('/subjects')
def subjects_list():
    conn = get_db()
    rows = conn.execute("""
        SELECT s.*, t.first_name||' '||t.last_name as teacher_name
        FROM subjects s LEFT JOIN teachers t ON t.id = s.teacher_id ORDER BY s.name
    """).fetchall()
    conn.close()

    rows_html = ""
    for r in rows:
        rows_html += f'''<tr>
            <td>{r['name']}</td><td>{r['code'] or '-'}</td><td>{r['coefficient']}</td>
            <td>{r['teacher_name'] or '-'}</td><td>{r['description'] or '-'}</td>
            <td>
              <a href="/subjects/edit/{r['id']}" class="btn btn-warning btn-sm">✏️</a>
              <a href="/subjects/delete/{r['id']}" class="btn btn-danger btn-sm" onclick="return confirm('Supprimer?')">🗑️</a>
            </td></tr>'''

    content = f'''
    <div class="toolbar">
      <div></div>
      <a href="/export/subjects" class="btn btn-success btn-sm">📊 Excel</a>
      <button class="btn btn-primary" onclick="openModal('modal-subject')">+ Nouvelle matière</button>
    </div>
    <div class="table-wrap"><table>
      <thead><tr><th>Nom</th><th>Code</th><th>Coef.</th><th>Enseignant</th><th>Description</th><th>Actions</th></tr></thead>
      <tbody>{rows_html}</tbody>
    </table></div>'''

    conn2 = get_db()
    teachers = conn2.execute("SELECT id, first_name, last_name FROM teachers ORDER BY last_name").fetchall()
    conn2.close()
    t_opts = '<option value="">-- Aucun --</option>' + "".join(
        f'<option value="{t["id"]}">{t["first_name"]} {t["last_name"]}</option>' for t in teachers
    )

    modal = f'''
    <div class="modal-overlay" id="modal-subject">
      <div class="modal">
        <button class="close-btn" onclick="closeModal('modal-subject')">&times;</button>
        <h2>Nouvelle matière</h2>
        <form method="post" action="/subjects/add">
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
            <div class="form-group"><label>Nom</label><input name="name" required></div>
            <div class="form-group"><label>Code</label><input name="code"></div>
            <div class="form-group"><label>Coefficient</label><input type="number" name="coefficient" value="1" min="1"></div>
            <div class="form-group"><label>Enseignant</label><select name="teacher_id" class="form-select">{t_opts}</select></div>
          </div>
          <div class="form-group"><label>Description</label><textarea name="description" rows="2"></textarea></div>
          <div style="margin-top:14px;display:flex;gap:8px;justify-content:flex-end;">
            <button type="button" class="btn btn-secondary" onclick="closeModal('modal-subject')">Annuler</button>
            <button type="submit" class="btn btn-primary">Enregistrer</button>
          </div>
        </form>
      </div>
    </div>'''
    return render("subjects", "Matières", content, modal)


@app.route('/subjects/add', methods=['POST'])
def subjects_add():
    try:
        conn = get_db()
        conn.execute("INSERT INTO subjects (name,code,coefficient,teacher_id,description) VALUES (?,?,?,?,?)",
            (request.form.get('name'), request.form.get('code'),
             int(request.form.get('coefficient') or 1),
             int(request.form.get('teacher_id') or 0) or None,
             request.form.get('description')))
        conn.commit()
        conn.close()
        flash("Matière ajoutée", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('subjects_list'))


@app.route('/subjects/edit/<int:sid>', methods=['GET', 'POST'])
def subjects_edit(sid):
    conn = get_db()
    s = conn.execute("SELECT * FROM subjects WHERE id=?", (sid,)).fetchone()
    if request.method == 'POST':
        try:
            conn.execute("UPDATE subjects SET name=?,code=?,coefficient=?,teacher_id=?,description=? WHERE id=?",
                (request.form.get('name'), request.form.get('code'),
                 int(request.form.get('coefficient') or 1),
                 int(request.form.get('teacher_id') or 0) or None,
                 request.form.get('description'), sid))
            conn.commit()
            conn.close()
            flash("Matière modifiée", "success")
            return redirect(url_for('subjects_list'))
        except Exception as e:
            flash(f"Erreur: {e}", "error")
    teachers = conn.execute("SELECT id, first_name, last_name FROM teachers ORDER BY last_name").fetchall()
    conn.close()
    t_opts = '<option value="">-- Aucun --</option>' + "".join(
        f'<option value="{t["id"]}" {"selected" if t["id"]==s["teacher_id"] else ""}>{t["first_name"]} {t["last_name"]}</option>' for t in teachers
    )
    content = f'''
    <div class="card" style="max-width:600px;">
      <h2 style="margin-bottom:16px;">Modifier la matière</h2>
      <form method="post">
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
          <div class="form-group"><label>Nom</label><input name="name" value="{s['name']}" required></div>
          <div class="form-group"><label>Code</label><input name="code" value="{s['code'] or ''}"></div>
          <div class="form-group"><label>Coefficient</label><input type="number" name="coefficient" value="{s['coefficient']}" min="1"></div>
          <div class="form-group"><label>Enseignant</label><select name="teacher_id" class="form-select">{t_opts}</select></div>
        </div>
        <div class="form-group"><label>Description</label><textarea name="description" rows="2">{s['description'] or ''}</textarea></div>
        <div style="margin-top:14px;display:flex;gap:8px;">
          <button type="submit" class="btn btn-success">Sauvegarder</button>
          <a href="/subjects" class="btn btn-secondary">Annuler</a>
        </div>
      </form>
    </div>'''
    return render("subjects", "Modifier matière", content)


@app.route('/subjects/delete/<int:sid>')
def subjects_delete(sid):
    try:
        conn = get_db()
        conn.execute("DELETE FROM subjects WHERE id=?", (sid,))
        conn.commit()
        conn.close()
        flash("Matière supprimée", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('subjects_list'))


# ────────────── GRADES ──────────────
@app.route('/grades')
def grades_list():
    conn = get_db()
    cycle = session.get('cycle', 'primaire')
    f_class = request.args.get('class_id', '')
    f_term = request.args.get('term', '')

    sql = """SELECT g.id, g.score, g.max_score, g.term, g.exam_type, g.date, g.comment,
             s.first_name||' '||s.last_name as student_name, s.matricule,
             sub.name as subject_name, sub.coefficient, c.name as class_name
             FROM grades g
             JOIN students s ON s.id = g.student_id
             JOIN subjects sub ON sub.id = g.subject_id
             LEFT JOIN classes c ON c.id = s.class_id WHERE s.cycle=?"""
    params = [cycle]
    if f_class:
        sql += " AND s.class_id = ?"
        params.append(int(f_class))
    if f_term:
        sql += " AND g.term = ?"
        params.append(f_term)
    sql += " ORDER BY s.last_name, sub.name"
    rows = conn.execute(sql, params).fetchall()

    classes = conn.execute("SELECT id, name FROM classes WHERE cycle=? ORDER BY name", (cycle,)).fetchall()
    conn.close()

    class_opts = '<option value="">Toutes</option>' + "".join(
        f'<option value="{c["id"]}" {"selected" if str(c["id"])==f_class else ""}>{c["name"]}</option>' for c in classes
    )
    term_opts = '<option value="">Tous</option>' + "".join(
        f'<option value="{t}" {"selected" if t==f_term else ""}>{t}</option>'
        for t in ["1er Trimestre", "2ème Trimestre", "3ème Trimestre"]
    )

    rows_html = ""
    for r in rows:
        pct = round(r['score'] / r['max_score'] * 100, 1) if r['max_score'] else 0
        pct_cls = "badge-success" if pct >= 50 else "badge-danger"
        rows_html += f'''<tr>
            <td>{r['matricule']}</td><td>{r['student_name']}</td><td>{r['class_name'] or '-'}</td>
            <td>{r['subject_name']}</td><td>{r['coefficient']}</td>
            <td>{r['term']}</td><td>{r['exam_type'] or '-'}</td>
            <td>{r['score']}/{r['max_score']}</td>
            <td><span class="badge {pct_cls}">{pct}%</span></td>
            <td>
              <a href="/grades/edit/{r['id']}" class="btn btn-warning btn-sm">✏️</a>
              <a href="/grades/delete/{r['id']}" class="btn btn-danger btn-sm" onclick="return confirm('Supprimer?')">🗑️</a>
            </td></tr>'''

    # Build student select and subject select for the modal
    conn2 = get_db()
    students = conn2.execute("SELECT id, matricule, first_name, last_name FROM students WHERE cycle=? ORDER BY last_name", (cycle,)).fetchall()
    subjects = conn2.execute("SELECT id, name FROM subjects ORDER BY name").fetchall()
    conn2.close()
    s_opts = "".join(f'<option value="{s["id"]}">{s["matricule"]} - {s["first_name"]} {s["last_name"]}</option>' for s in students)
    sub_opts = "".join(f'<option value="{sub["id"]}">{sub["name"]}</option>' for sub in subjects)

    content = f'''
    <div class="toolbar">
      <form method="get" style="display:flex;gap:8px;flex-wrap:wrap;">
        <select name="class_id" onchange="this.form.submit()">{class_opts}</select>
        <select name="term" onchange="this.form.submit()">{term_opts}</select>
      </form>
      <a href="/export/grades" class="btn btn-success btn-sm">📊 Excel</a>
      <button class="btn btn-primary" onclick="openModal('modal-grade')">+ Nouvelle note</button>
    </div>
    <div class="table-wrap"><table>
      <thead><tr><th>Matricule</th><th>Élève</th><th>Classe</th><th>Matière</th><th>Coef.</th><th>Trimestre</th><th>Type</th><th>Note</th><th>%</th><th>Actions</th></tr></thead>
      <tbody>{rows_html}</tbody>
    </table></div>'''

    modal = f'''
    <div class="modal-overlay" id="modal-grade">
      <div class="modal">
        <button class="close-btn" onclick="closeModal('modal-grade')">&times;</button>
        <h2>Nouvelle note</h2>
        <form method="post" action="/grades/add">
          <div class="form-group"><label>Élève</label><select name="student_id" class="form-select" required>{s_opts}</select></div>
          <div class="form-group"><label>Matière</label><select name="subject_id" class="form-select" required>{sub_opts}</select></div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
            <div class="form-group"><label>Trimestre</label><select name="term" class="form-select"><option>1er Trimestre</option><option>2ème Trimestre</option><option>3ème Trimestre</option></select></div>
            <div class="form-group"><label>Type d'examen</label><select name="exam_type" class="form-select"><option>Devoir</option><option>Examen</option><option>Contrôle</option></select></div>
            <div class="form-group"><label>Note (/20)</label><input type="number" name="score" step="0.5" min="0" max="20" required></div>
            <div class="form-group"><label>Note maximale</label><input type="number" name="max_score" value="20" min="1" required></div>
            <div class="form-group"><label>Date</label><input type="date" name="date"></div>
          </div>
          <div class="form-group"><label>Commentaire</label><input name="comment"></div>
          <div style="margin-top:14px;display:flex;gap:8px;justify-content:flex-end;">
            <button type="button" class="btn btn-secondary" onclick="closeModal('modal-grade')">Annuler</button>
            <button type="submit" class="btn btn-primary">Enregistrer</button>
          </div>
        </form>
      </div>
    </div>'''
    return render("grades", "Notes", content, modal)


@app.route('/grades/add', methods=['POST'])
def grades_add():
    try:
        conn = get_db()
        conn.execute("""INSERT INTO grades (student_id,subject_id,term,exam_type,score,max_score,date,comment)
            VALUES (?,?,?,?,?,?,?,?)""",
            (int(request.form.get('student_id')), int(request.form.get('subject_id')),
             request.form.get('term'), request.form.get('exam_type'),
             float(request.form.get('score')), float(request.form.get('max_score')),
             request.form.get('date') or date.today().isoformat(), request.form.get('comment')))
        conn.commit()
        conn.close()
        flash("Note ajoutée", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('grades_list'))


@app.route('/grades/edit/<int:gid>', methods=['GET', 'POST'])
def grades_edit(gid):
    conn = get_db()
    g = conn.execute("SELECT * FROM grades WHERE id=?", (gid,)).fetchone()
    if request.method == 'POST':
        try:
            conn.execute("""UPDATE grades SET student_id=?,subject_id=?,term=?,exam_type=?,score=?,max_score=?,date=?,comment=? WHERE id=?""",
                (int(request.form.get('student_id')), int(request.form.get('subject_id')),
                 request.form.get('term'), request.form.get('exam_type'),
                 float(request.form.get('score')), float(request.form.get('max_score')),
                 request.form.get('date'), request.form.get('comment'), gid))
            conn.commit()
            conn.close()
            flash("Note modifiée", "success")
            return redirect(url_for('grades_list'))
        except Exception as e:
            flash(f"Erreur: {e}", "error")
    students = conn.execute("SELECT id, matricule, first_name, last_name FROM students ORDER BY last_name").fetchall()
    subjects = conn.execute("SELECT id, name FROM subjects ORDER BY name").fetchall()
    conn.close()
    s_opts = "".join(f'<option value="{s["id"]}" {"selected" if s["id"]==g["student_id"] else ""}>{s["matricule"]} - {s["first_name"]} {s["last_name"]}</option>' for s in students)
    sub_opts = "".join(f'<option value="{sub["id"]}" {"selected" if sub["id"]==g["subject_id"] else ""}>{sub["name"]}</option>' for sub in subjects)

    content = f'''
    <div class="card" style="max-width:600px;">
      <h2 style="margin-bottom:16px;">Modifier la note</h2>
      <form method="post">
        <div class="form-group"><label>Élève</label><select name="student_id" class="form-select">{s_opts}</select></div>
        <div class="form-group"><label>Matière</label><select name="subject_id" class="form-select">{sub_opts}</select></div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
          <div class="form-group"><label>Trimestre</label><select name="term" class="form-select">
            {"<option selected>1er Trimestre</option>" if "1er" in (g["term"] or "") else "<option>1er Trimestre</option>"}
            {"<option selected>2ème Trimestre</option>" if "2ème" in (g["term"] or "") else "<option>2ème Trimestre</option>"}
            {"<option selected>3ème Trimestre</option>" if "3ème" in (g["term"] or "") else "<option>3ème Trimestre</option>"}
          </select></div>
          <div class="form-group"><label>Type</label><select name="exam_type" class="form-select">
            {"<option selected>Devoir</option>" if g["exam_type"]=="Devoir" else "<option>Devoir</option>"}
            {"<option selected>Examen</option>" if g["exam_type"]=="Examen" else "<option>Examen</option>"}
            {"<option selected>Contrôle</option>" if g["exam_type"]=="Contrôle" else "<option>Contrôle</option>"}
          </select></div>
          <div class="form-group"><label>Note</label><input type="number" name="score" value="{g['score']}" step="0.5" required></div>
          <div class="form-group"><label>Max</label><input type="number" name="max_score" value="{g['max_score']}" required></div>
          <div class="form-group"><label>Date</label><input type="date" name="date" value="{g['date'] or ''}"></div>
        </div>
        <div class="form-group"><label>Commentaire</label><input name="comment" value="{g['comment'] or ''}"></div>
        <div style="margin-top:14px;display:flex;gap:8px;">
          <button type="submit" class="btn btn-success">Sauvegarder</button>
          <a href="/grades" class="btn btn-secondary">Annuler</a>
        </div>
      </form>
    </div>'''
    return render("grades", "Modifier note", content)


@app.route('/grades/delete/<int:gid>')
def grades_delete(gid):
    try:
        conn = get_db()
        conn.execute("DELETE FROM grades WHERE id=?", (gid,))
        conn.commit()
        conn.close()
        flash("Note supprimée", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('grades_list'))


# ────────────── ATTENDANCE ──────────────
@app.route('/attendance')
def attendance_list():
    conn = get_db()
    cycle = session.get('cycle', 'primaire')
    f_date = request.args.get('date', date.today().isoformat())
    f_class = request.args.get('class_id', '')

    sql = """SELECT a.id, a.student_id, a.date, a.status, a.comment,
             s.first_name||' '||s.last_name as student_name, s.matricule, c.name as class_name
             FROM attendance a
             JOIN students s ON s.id = a.student_id
             LEFT JOIN classes c ON c.id = s.class_id WHERE s.cycle=?"""
    params = [cycle]
    if f_date:
        sql += " AND a.date = ?"
        params.append(f_date)
    if f_class:
        sql += " AND s.class_id = ?"
        params.append(int(f_class))
    sql += " ORDER BY c.name, s.last_name"
    rows = conn.execute(sql, params).fetchall()

    classes = conn.execute("SELECT id, name FROM classes WHERE cycle=? ORDER BY name", (cycle,)).fetchall()
    conn.close()

    class_opts = '<option value="">Toutes</option>' + "".join(
        f'<option value="{c["id"]}" {"selected" if str(c["id"])==f_class else ""}>{c["name"]}</option>' for c in classes
    )

    rows_html = ""
    for r in rows:
        rows_html += f'''<tr>
            <td>{r['matricule']}</td><td>{r['student_name']}</td><td>{r['class_name'] or '-'}</td>
            <td>{r['date']}</td><td>{badge(r['status'])}</td><td>{r['comment'] or '-'}</td>
            <td>
              <a href="/attendance/edit/{r['id']}" class="btn btn-warning btn-sm">✏️</a>
              <a href="/attendance/delete/{r['id']}" class="btn btn-danger btn-sm" onclick="return confirm('Supprimer?')">🗑️</a>
            </td></tr>'''

    conn2 = get_db()
    students = conn2.execute("SELECT id, matricule, first_name, last_name FROM students WHERE cycle=? ORDER BY last_name", (cycle,)).fetchall()
    conn2.close()
    s_opts = "".join(f'<option value="{s["id"]}">{s["matricule"]} - {s["first_name"]} {s["last_name"]}</option>' for s in students)

    content = f'''
    <div class="toolbar">
      <form method="get" style="display:flex;gap:8px;flex-wrap:wrap;">
        <input type="date" name="date" value="{f_date}" onchange="this.form.submit()">
        <select name="class_id" onchange="this.form.submit()">{class_opts}</select>
      </form>
      <a href="/export/attendance" class="btn btn-success btn-sm">📊 Excel</a>
      <button class="btn btn-primary" onclick="openModal('modal-att')">+ Marquer présence</button>
    </div>
    <div class="table-wrap"><table>
      <thead><tr><th>Matricule</th><th>Élève</th><th>Classe</th><th>Date</th><th>Statut</th><th>Commentaire</th><th>Actions</th></tr></thead>
      <tbody>{rows_html}</tbody>
    </table></div>'''

    modal = f'''
    <div class="modal-overlay" id="modal-att">
      <div class="modal">
        <button class="close-btn" onclick="closeModal('modal-att')">&times;</button>
        <h2>Marquer présence</h2>
        <form method="post" action="/attendance/add">
          <div class="form-group"><label>Élève</label><select name="student_id" class="form-select" required>{s_opts}</select></div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
            <div class="form-group"><label>Date</label><input type="date" name="date" value="{date.today().isoformat()}" required></div>
            <div class="form-group"><label>Statut</label><select name="status" class="form-select"><option value="présent">Présent</option><option value="absent">Absent</option><option value="retard">Retard</option></select></div>
          </div>
          <div class="form-group"><label>Commentaire</label><input name="comment"></div>
          <div style="margin-top:14px;display:flex;gap:8px;justify-content:flex-end;">
            <button type="button" class="btn btn-secondary" onclick="closeModal('modal-att')">Annuler</button>
            <button type="submit" class="btn btn-primary">Enregistrer</button>
          </div>
        </form>
      </div>
    </div>'''
    return render("attendance", "Présences", content, modal)


@app.route('/attendance/add', methods=['POST'])
def attendance_add():
    try:
        conn = get_db()
        conn.execute("INSERT INTO attendance (student_id,date,status,comment) VALUES (?,?,?,?)",
            (int(request.form.get('student_id')), request.form.get('date'),
             request.form.get('status'), request.form.get('comment')))
        conn.commit()
        conn.close()
        flash("Présence enregistrée", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('attendance_list'))


@app.route('/attendance/edit/<int:aid>', methods=['GET', 'POST'])
def attendance_edit(aid):
    conn = get_db()
    a = conn.execute("SELECT * FROM attendance WHERE id=?", (aid,)).fetchone()
    if request.method == 'POST':
        try:
            conn.execute("UPDATE attendance SET student_id=?,date=?,status=?,comment=? WHERE id=?",
                (int(request.form.get('student_id')), request.form.get('date'),
                 request.form.get('status'), request.form.get('comment'), aid))
            conn.commit()
            conn.close()
            flash("Présence modifiée", "success")
            return redirect(url_for('attendance_list'))
        except Exception as e:
            flash(f"Erreur: {e}", "error")
    students = conn.execute("SELECT id, matricule, first_name, last_name FROM students ORDER BY last_name").fetchall()
    conn.close()
    s_opts = "".join(f'<option value="{s["id"]}" {"selected" if s["id"]==a["student_id"] else ""}>{s["matricule"]} - {s["first_name"]} {s["last_name"]}</option>' for s in students)
    content = f'''
    <div class="card" style="max-width:600px;">
      <h2 style="margin-bottom:16px;">Modifier la présence</h2>
      <form method="post">
        <div class="form-group"><label>Élève</label><select name="student_id" class="form-select">{s_opts}</select></div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
          <div class="form-group"><label>Date</label><input type="date" name="date" value="{a['date']}" required></div>
          <div class="form-group"><label>Statut</label><select name="status" class="form-select">
            {"<option selected>présent</option>" if a["status"]=="présent" else "<option>présent</option>"}
            {"<option selected>absent</option>" if a["status"]=="absent" else "<option>absent</option>"}
            {"<option selected>retard</option>" if a["status"]=="retard" else "<option>retard</option>"}
          </select></div>
        </div>
        <div class="form-group"><label>Commentaire</label><input name="comment" value="{a['comment'] or ''}"></div>
        <div style="margin-top:14px;display:flex;gap:8px;">
          <button type="submit" class="btn btn-success">Sauvegarder</button>
          <a href="/attendance" class="btn btn-secondary">Annuler</a>
        </div>
      </form>
    </div>'''
    return render("attendance", "Modifier présence", content)


@app.route('/attendance/delete/<int:aid>')
def attendance_delete(aid):
    try:
        conn = get_db()
        conn.execute("DELETE FROM attendance WHERE id=?", (aid,))
        conn.commit()
        conn.close()
        flash("Présence supprimée", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('attendance_list'))


# ────────────── TIMETABLE ──────────────
@app.route('/timetable')
def timetable_list():
    conn = get_db()
    cycle = session.get('cycle', 'primaire')
    f_class = request.args.get('class_id', '')
    classes = conn.execute("SELECT id, name FROM classes WHERE cycle=? ORDER BY name", (cycle,)).fetchall()

    sql = """SELECT tt.*, c.name as class_name, sub.name as subject_name, t.first_name||' '||t.last_name as teacher_name
             FROM timetable tt
             JOIN classes c ON c.id = tt.class_id
             JOIN subjects sub ON sub.id = tt.subject_id
             LEFT JOIN teachers t ON t.id = sub.teacher_id WHERE c.cycle=?"""
    params = [cycle]
    if f_class:
        sql += " AND tt.class_id = ?"
        params.append(int(f_class))
    sql += " ORDER BY tt.class_id, CASE tt.day WHEN 'Lundi' THEN 1 WHEN 'Mardi' THEN 2 WHEN 'Mercredi' THEN 3 WHEN 'Jeudi' THEN 4 WHEN 'Vendredi' THEN 5 WHEN 'Samedi' THEN 6 ELSE 7 END, tt.time_start"
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    class_opts = '<option value="">Toutes</option>' + "".join(
        f'<option value="{c["id"]}" {"selected" if str(c["id"])==f_class else ""}>{c["name"]}</option>' for c in classes
    )

    rows_html = ""
    for r in rows:
        rows_html += f'''<tr>
            <td>{r['class_name']}</td><td>{r['day']}</td>
            <td>{r['time_start']} - {r['time_end']}</td>
            <td>{r['subject_name']}</td><td>{r['room'] or '-'}</td><td>{r['teacher_name'] or '-'}</td>
            <td>
              <a href="/timetable/edit/{r['id']}" class="btn btn-warning btn-sm">✏️</a>
              <a href="/timetable/delete/{r['id']}" class="btn btn-danger btn-sm" onclick="return confirm('Supprimer?')">🗑️</a>
            </td></tr>'''

    # Modal selects
    conn2 = get_db()
    classes2 = conn2.execute("SELECT id, name FROM classes WHERE cycle=? ORDER BY name", (cycle,)).fetchall()
    subjects = conn2.execute("SELECT id, name FROM subjects ORDER BY name").fetchall()
    conn2.close()
    c_opts = "".join(f'<option value="{c["id"]}">{c["name"]}</option>' for c in classes2)
    sub_opts = "".join(f'<option value="{s["id"]}">{s["name"]}</option>' for s in subjects)
    day_opts = "".join(f'<option>{d}</option>' for d in ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"])

    content = f'''
    <div class="toolbar">
      <form method="get" style="display:flex;gap:8px;">
        <select name="class_id" onchange="this.form.submit()">{class_opts}</select>
      </form>
      <a href="/export/timetable" class="btn btn-success btn-sm">📊 Excel</a>
      <button class="btn btn-primary" onclick="openModal('modal-tt')">+ Nouveau créneau</button>
    </div>
    <div class="table-wrap"><table>
      <thead><tr><th>Classe</th><th>Jour</th><th>Horaire</th><th>Matière</th><th>Salle</th><th>Enseignant</th><th>Actions</th></tr></thead>
      <tbody>{rows_html}</tbody>
    </table></div>'''

    modal = f'''
    <div class="modal-overlay" id="modal-tt">
      <div class="modal">
        <button class="close-btn" onclick="closeModal('modal-tt')">&times;</button>
        <h2>Nouveau créneau horaire</h2>
        <form method="post" action="/timetable/add">
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
            <div class="form-group"><label>Classe</label><select name="class_id" class="form-select" required>{c_opts}</select></div>
            <div class="form-group"><label>Matière</label><select name="subject_id" class="form-select" required>{sub_opts}</select></div>
            <div class="form-group"><label>Jour</label><select name="day" class="form-select">{day_opts}</select></div>
            <div class="form-group"><label>Salle</label><input name="room"></div>
            <div class="form-group"><label>Début</label><input type="time" name="time_start" required></div>
            <div class="form-group"><label>Fin</label><input type="time" name="time_end" required></div>
          </div>
          <div style="margin-top:14px;display:flex;gap:8px;justify-content:flex-end;">
            <button type="button" class="btn btn-secondary" onclick="closeModal('modal-tt')">Annuler</button>
            <button type="submit" class="btn btn-primary">Enregistrer</button>
          </div>
        </form>
      </div>
    </div>'''
    return render("timetable", "Emploi du temps", content, modal)


@app.route('/timetable/add', methods=['POST'])
def timetable_add():
    try:
        conn = get_db()
        conn.execute("INSERT INTO timetable (class_id,subject_id,day,time_start,time_end,room) VALUES (?,?,?,?,?,?)",
            (int(request.form.get('class_id')), int(request.form.get('subject_id')),
             request.form.get('day'), request.form.get('time_start'),
             request.form.get('time_end'), request.form.get('room')))
        conn.commit()
        conn.close()
        flash("Créneau ajouté", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('timetable_list'))


@app.route('/timetable/edit/<int:tid>', methods=['GET', 'POST'])
def timetable_edit(tid):
    conn = get_db()
    t = conn.execute("SELECT * FROM timetable WHERE id=?", (tid,)).fetchone()
    if request.method == 'POST':
        try:
            conn.execute("UPDATE timetable SET class_id=?,subject_id=?,day=?,time_start=?,time_end=?,room=? WHERE id=?",
                (int(request.form.get('class_id')), int(request.form.get('subject_id')),
                 request.form.get('day'), request.form.get('time_start'),
                 request.form.get('time_end'), request.form.get('room'), tid))
            conn.commit()
            conn.close()
            flash("Créneau modifié", "success")
            return redirect(url_for('timetable_list'))
        except Exception as e:
            flash(f"Erreur: {e}", "error")
    classes = conn.execute("SELECT id, name FROM classes ORDER BY name").fetchall()
    subjects = conn.execute("SELECT id, name FROM subjects ORDER BY name").fetchall()
    conn.close()
    c_opts = "".join(f'<option value="{c["id"]}" {"selected" if c["id"]==t["class_id"] else ""}>{c["name"]}</option>' for c in classes)
    sub_opts = "".join(f'<option value="{s["id"]}" {"selected" if s["id"]==t["subject_id"] else ""}>{s["name"]}</option>' for s in subjects)
    day_opts = "".join(f'<option {"selected" if d==t["day"] else ""}>{d}</option>' for d in ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"])

    content = f'''
    <div class="card" style="max-width:600px;">
      <h2 style="margin-bottom:16px;">Modifier le créneau</h2>
      <form method="post">
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
          <div class="form-group"><label>Classe</label><select name="class_id" class="form-select">{c_opts}</select></div>
          <div class="form-group"><label>Matière</label><select name="subject_id" class="form-select">{sub_opts}</select></div>
          <div class="form-group"><label>Jour</label><select name="day" class="form-select">{day_opts}</select></div>
          <div class="form-group"><label>Salle</label><input name="room" value="{t['room'] or ''}"></div>
          <div class="form-group"><label>Début</label><input type="time" name="time_start" value="{t['time_start']}" required></div>
          <div class="form-group"><label>Fin</label><input type="time" name="time_end" value="{t['time_end']}" required></div>
        </div>
        <div style="margin-top:14px;display:flex;gap:8px;">
          <button type="submit" class="btn btn-success">Sauvegarder</button>
          <a href="/timetable" class="btn btn-secondary">Annuler</a>
        </div>
      </form>
    </div>'''
    return render("timetable", "Modifier créneau", content)


@app.route('/timetable/delete/<int:tid>')
def timetable_delete(tid):
    try:
        conn = get_db()
        conn.execute("DELETE FROM timetable WHERE id=?", (tid,))
        conn.commit()
        conn.close()
        flash("Créneau supprimé", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('timetable_list'))


# ────────────── FEES ──────────────
@app.route('/fees')
def fees_list():
    conn = get_db()
    cycle = session.get('cycle', 'primaire')
    f_status = request.args.get('status', '')

    sql = """SELECT f.*, s.first_name||' '||s.last_name as student_name, s.matricule
             FROM fees f JOIN students s ON s.id = f.student_id WHERE s.cycle=?"""
    params = [cycle]
    if f_status:
        sql += " AND f.status = ?"
        params.append(f_status)
    sql += " ORDER BY s.last_name"
    rows = conn.execute(sql, params).fetchall()

    # Totals
    totals = conn.execute("SELECT SUM(f.amount) as total, SUM(f.amount_paid) as paid FROM fees f JOIN students s ON s.id=f.student_id WHERE s.cycle=?", (cycle,)).fetchone()
    conn.close()

    total_amount = totals['total'] or 0
    total_paid = totals['paid'] or 0
    remaining = total_amount - total_paid

    status_opts = '<option value="">Tous</option>' + "".join(
        f'<option value="{st}" {"selected" if st==f_status else ""}>{st}</option>'
        for st in ["payé", "partiel", "impayé"]
    )

    rows_html = ""
    for r in rows:
        rows_html += f'''<tr>
            <td>{r['matricule']}</td><td>{r['student_name']}</td>
            <td>{r['fee_type'] or '-'}</td>
            <td>{r['amount']:,.0f} CDF</td><td>{r['amount_paid']:,.0f} CDF</td>
            <td>{r['due_date'] or '-'}</td><td>{r['payment_date'] or '-'}</td>
            <td>{r['receipt_number'] or '-'}</td><td>{badge(r['status'])}</td>
            <td>
              <a href="/fees/receipt/{r['id']}" class="btn btn-success btn-sm" target="_blank" title="Imprimer le reçu" style="font-size:15px;padding:6px 14px">🧾 <b>Reçu</b></a>
              <a href="/fees/edit/{r['id']}" class="btn btn-warning btn-sm">✏️</a>
              <a href="/fees/delete/{r['id']}" class="btn btn-danger btn-sm" onclick="return confirm('Supprimer?')">🗑️</a>
            </td></tr>'''

    conn2 = get_db()
    students = conn2.execute("SELECT id, matricule, first_name, last_name FROM students WHERE cycle=? ORDER BY last_name", (cycle,)).fetchall()
    conn2.close()
    s_opts = "".join(f'<option value="{s["id"]}">{s["matricule"]} - {s["first_name"]} {s["last_name"]}</option>' for s in students)

    content = f'''
    <div class="cards">
      <div class="card card-green"><span class="card-icon">💰</span><div class="card-label">Total encaissé</div><div class="card-value">{total_paid:,.0f} CDF</div></div>
      <div class="card card-red"><span class="card-icon">📊</span><div class="card-label">Total restant</div><div class="card-value">{remaining:,.0f} CDF</div></div>
      <div class="card card-blue"><span class="card-icon">📋</span><div class="card-label">Total facturé</div><div class="card-value">{total_amount:,.0f} CDF</div></div>
    </div>
    <div class="toolbar">
      <form method="get" style="display:flex;gap:8px;">
        <select name="status" onchange="this.form.submit()">{status_opts}</select>
      </form>
      <a href="/export/fees" class="btn btn-success btn-sm">📊 Excel</a>
      <button class="btn btn-primary" onclick="openModal('modal-fee')">+ Nouveau frais</button>
    </div>
    <div class="table-wrap"><table>
      <thead><tr><th>Matricule</th><th>Élève</th><th>Type</th><th>Montant</th><th>Payé</th><th>Échéance</th><th>Date paiement</th><th>Reçu</th><th>Statut</th><th>Actions</th></tr></thead>
      <tbody>{rows_html}</tbody>
    </table></div>'''

    modal = f'''
    <div class="modal-overlay" id="modal-fee">
      <div class="modal">
        <button class="close-btn" onclick="closeModal('modal-fee')">&times;</button>
        <h2>Nouveau frais scolaire</h2>
        <form method="post" action="/fees/add">
          <div class="form-group"><label>Élève</label><select name="student_id" class="form-select" required>{s_opts}</select></div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
            <div class="form-group"><label>Type de frais</label><input name="fee_type" value="Inscription"></div>
            <div class="form-group"><label>Montant (CDF)</label><input type="number" name="amount" value="150000" required></div>
            <div class="form-group"><label>Montant payé (CDF)</label><input type="number" name="amount_paid" value="0"></div>
            <div class="form-group"><label>Date d'échéance</label><input type="date" name="due_date"></div>
            <div class="form-group"><label>Date de paiement</label><input type="date" name="payment_date"></div>
            <div class="form-group"><label>N° Reçu</label><input name="receipt_number"></div>
            <div class="form-group"><label>Statut</label><select name="status" class="form-select"><option value="impayé">Impayé</option><option value="partiel">Partiel</option><option value="payé">Payé</option></select></div>
          </div>
          <div class="form-group"><label>Commentaire</label><input name="comment"></div>
          <div style="margin-top:14px;display:flex;gap:8px;justify-content:flex-end;">
            <button type="button" class="btn btn-secondary" onclick="closeModal('modal-fee')">Annuler</button>
            <button type="submit" class="btn btn-primary">Enregistrer</button>
          </div>
        </form>
      </div>
    </div>'''
    return render("fees", "Frais scolaires", content, modal)


@app.route('/fees/add', methods=['POST'])
def fees_add():
    try:
        conn = get_db()
        cur = conn.execute("""INSERT INTO fees (student_id,fee_type,amount,amount_paid,due_date,payment_date,status,receipt_number,comment)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            (int(request.form.get('student_id')), request.form.get('fee_type'),
             float(request.form.get('amount')), float(request.form.get('amount_paid')),
             request.form.get('due_date'), request.form.get('payment_date'),
             request.form.get('status'), request.form.get('receipt_number'),
             request.form.get('comment')))
        conn.commit()
        new_id = cur.lastrowid
        conn.close()
        # Si paiement effectué, rediriger vers le reçu directement
        amount_paid = float(request.form.get('amount_paid', 0))
        if amount_paid > 0:
            return redirect(url_for('fees_receipt', fid=new_id))
        flash("Frais ajouté", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('fees_list'))


@app.route('/fees/edit/<int:fid>', methods=['GET', 'POST'])
def fees_edit(fid):
    conn = get_db()
    f = conn.execute("SELECT * FROM fees WHERE id=?", (fid,)).fetchone()
    if request.method == 'POST':
        try:
            conn.execute("""UPDATE fees SET student_id=?,fee_type=?,amount=?,amount_paid=?,due_date=?,payment_date=?,status=?,receipt_number=?,comment=? WHERE id=?""",
                (int(request.form.get('student_id')), request.form.get('fee_type'),
                 float(request.form.get('amount')), float(request.form.get('amount_paid')),
                 request.form.get('due_date'), request.form.get('payment_date'),
                 request.form.get('status'), request.form.get('receipt_number'),
                 request.form.get('comment'), fid))
            conn.commit()
            conn.close()
            flash("Frais modifié", "success")
            # Si paiement effectué, rediriger vers le reçu directement
            amount_paid = float(request.form.get('amount_paid', 0))
            if amount_paid > 0:
                return redirect(url_for('fees_receipt', fid=fid))
            return redirect(url_for('fees_list'))
        except Exception as e:
            flash(f"Erreur: {e}", "error")
    students = conn.execute("SELECT id, matricule, first_name, last_name FROM students ORDER BY last_name").fetchall()
    conn.close()
    s_opts = "".join(f'<option value="{s["id"]}" {"selected" if s["id"]==f["student_id"] else ""}>{s["matricule"]} - {s["first_name"]} {s["last_name"]}</option>' for s in students)

    content = f'''
    <div class="card" style="max-width:600px;">
      <h2 style="margin-bottom:16px;">Modifier les frais</h2>
      <form method="post">
        <div class="form-group"><label>Élève</label><select name="student_id" class="form-select">{s_opts}</select></div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 12px;">
          <div class="form-group"><label>Type</label><input name="fee_type" value="{f['fee_type'] or ''}"></div>
          <div class="form-group"><label>Montant (CDF)</label><input type="number" name="amount" value="{f['amount']}" required></div>
          <div class="form-group"><label>Payé (CDF)</label><input type="number" name="amount_paid" value="{f['amount_paid']}"></div>
          <div class="form-group"><label>Échéance</label><input type="date" name="due_date" value="{f['due_date'] or ''}"></div>
          <div class="form-group"><label>Date paiement</label><input type="date" name="payment_date" value="{f['payment_date'] or ''}"></div>
          <div class="form-group"><label>N° Reçu</label><input name="receipt_number" value="{f['receipt_number'] or ''}"></div>
          <div class="form-group"><label>Statut</label><select name="status" class="form-select">
            {"<option selected>impayé</option>" if f["status"]=="impayé" else "<option>impayé</option>"}
            {"<option selected>partiel</option>" if f["status"]=="partiel" else "<option>partiel</option>"}
            {"<option selected>payé</option>" if f["status"]=="payé" else "<option>payé</option>"}
          </select></div>
        </div>
        <div class="form-group"><label>Commentaire</label><input name="comment" value="{f['comment'] or ''}"></div>
        <div style="margin-top:14px;display:flex;gap:8px;">
          <button type="submit" class="btn btn-success">Sauvegarder</button>
          <a href="/fees" class="btn btn-secondary">Annuler</a>
        </div>
      </form>
    </div>'''
    return render("fees", "Modifier frais", content)


@app.route('/fees/delete/<int:fid>')
def fees_delete(fid):
    try:
        conn = get_db()
        conn.execute("DELETE FROM fees WHERE id=?", (fid,))
        conn.commit()
        conn.close()
        flash("Frais supprimé", "success")
    except Exception as e:
        flash(f"Erreur: {e}", "error")
    return redirect(url_for('fees_list'))


# ────────────── RECEIPT ──────────────
@app.route('/fees/receipt/<int:fid>')
def fees_receipt(fid):
    conn = get_db()
    f = conn.execute("""SELECT f.*, s.first_name||' '||s.last_name as student_name, s.matricule,
        s.class_id, c.name as class_name, s.parent_name, s.parent_phone
        FROM fees f JOIN students s ON s.id = f.student_id LEFT JOIN classes c ON c.id = s.class_id
        WHERE f.id=?""", (fid,)).fetchone()
    if not f:
        conn.close()
        return "Fee not found", 404
    receipt_num = f['receipt_number'] or f"REC-{fid:04d}"
    remaining = f['amount'] - f['amount_paid']
    status_label = {"payé": "PAYE EN TOTALITE", "partiel": "PAIEMENT PARTIEL", "impayé": "NON PAYE"}.get(f['status'], f['status'])
    conn.close()
    receipt_html = f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Recu {receipt_num}</title>
<style>
@page{{size:80mm auto;margin:8mm}}body{{font-family:'Segoe UI',Arial,sans-serif;max-width:320px;margin:0 auto;padding:15px;font-size:13px;color:#333}}
.center{{text-align:center}}.bold{{font-weight:700}}.line{{border-top:2px dashed #999;margin:10px 0}}
.header{{font-size:18px;font-weight:700;color:#1a1a2e;margin-bottom:2px}}.sub{{font-size:11px;color:#666}}
table{{width:100%;border-collapse:collapse;margin:10px 0}}td{{padding:4px 2px;font-size:12px}}
td:last-child{{text-align:right;font-weight:600}}.total-row td{{font-size:15px;font-weight:700;border-top:2px solid #333;padding-top:8px}}
.footer{{margin-top:15px;font-size:10px;color:#888;text-align:center}}
.stamp{{margin-top:30px;border:1px solid #333;border-radius:50%;width:80px;height:80px;display:inline-block;line-height:80px;text-align:center;font-size:10px;color:#1a1a2e}}
@media print{{body{{margin:0;padding:10px}}}}
</style></head>
<body>
<div class="center">
<div class="header">CongoSchool</div>
<div class="sub">Gestion Scolaire - Republique Democratique du Congo</div>
</div>
<div class="line"></div>
<div class="center"><div class="bold" style="font-size:14px;margin-bottom:3px">RECU DE PAIEMENT</div>
<div style="font-size:11px;color:#666">N {receipt_num} | Date: {f['payment_date'] or date.today().isoformat()}</div></div>
<div class="line"></div>
<table>
<tr><td>Eleve</td><td>{f['student_name']}</td></tr>
<tr><td>Matricule</td><td>{f['matricule']}</td></tr>
<tr><td>Classe</td><td>{f['class_name'] or '-'}</td></tr>
<tr><td>Parent</td><td>{f['parent_name'] or '-'}</td></tr>
</table>
<div class="line"></div>
<table>
<tr><td>Type de frais</td><td>{f['fee_type']}</td></tr>
<tr><td>Montant total</td><td>{f['amount']:,.0f} CDF</td></tr>
<tr><td>Montant paye</td><td style="color:#10b981">{f['amount_paid']:,.0f} CDF</td></tr>
<tr><td>Reste</td><td style="color:{"#ef4444" if remaining>0 else "#10b981"}">{remaining:,.0f} CDF</td></tr>
</table>
<div class="line"></div>
<div class="center" style="margin:10px 0">
<span style="display:inline-block;padding:4px 14px;border-radius:6px;font-size:12px;font-weight:700;
background:{"#d1fae5;color:#065f46" if f["status"]=="payé" else "#fef3c7;color:#92400e" if f["status"]=="partiel" else "#fee2e2;color:#991b1b"}">
{status_label}</span></div>
<div class="line"></div>
<div class="center">
<div class="footer">Ce recu fait foi de paiement.<br>Conservez-le precieusement.</div>
</div>
<div style="display:flex;justify-content:space-between;margin-top:25px;font-size:11px">
<div style="text-align:center"><div style="border-bottom:1px solid #333;width:100px;margin-bottom:3px"></div>Signature du parent</div>
<div style="text-align:center"><div style="border-bottom:1px solid #333;width:100px;margin-bottom:3px"></div>Cachet de l'etablissement</div>
</div>
<div style="margin-top:15px;text-align:center"><button onclick="window.print()" style="padding:8px 20px;border:none;border-radius:8px;background:#e94560;color:#fff;font-size:13px;cursor:pointer">Imprimer le recu</button></div>
</body></html>'''
    return receipt_html


# ────────────── EXCEL EXPORT ──────────────
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))

def make_export(headers, rows, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Donnees"
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
    for c in range(1, len(headers)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = max(16, max(len(str(h)) for h in headers)+4)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

EXPORT_BTN = '<a href="{url}" class="btn btn-success btn-sm" style="margin-left:8px">📊 Exporter Excel</a>'

@app.route('/export/students')
def export_students():
    cycle = session.get('cycle', 'primaire')
    conn = get_db()
    rows = conn.execute("""SELECT s.matricule, s.last_name, s.first_name, s.date_of_birth, s.gender,
        c.name, s.parent_name, s.parent_phone, s.address, s.email, s.enrollment_date, s.status
        FROM students s LEFT JOIN classes c ON c.id = s.class_id WHERE s.cycle=? ORDER BY s.last_name""", (cycle,)).fetchall()
    conn.close()
    data = [[r[0],r[1],r[2],r[3],r[4],r[5] or '',r[6] or '',r[7] or '',r[8] or '',r[9] or '',r[10],r[11]] for r in rows]
    return make_export(["Matricule","Nom","Prenom","Date naissance","Sexe","Classe","Parent","Telephone parent","Adresse","Email","Date inscription","Statut"], data, f"Eleves_{cycle.title()}_CongoSchool.xlsx")

@app.route('/export/teachers')
def export_teachers():
    conn = get_db()
    rows = conn.execute("SELECT first_name, last_name, phone, email, subject, address, hire_date, salary, status FROM teachers ORDER BY last_name").fetchall()
    conn.close()
    data = [[r[0],r[1],r[2] or '',r[3] or '',r[4] or '',r[5] or '',r[6],r[7],r[8]] for r in rows]
    return make_export(["Prenom","Nom","Telephone","Email","Matiere","Adresse","Date embauche","Salaire","Statut"], data, "Enseignants_CongoSchool.xlsx")

@app.route('/export/classes')
def export_classes():
    cycle = session.get('cycle', 'primaire')
    conn = get_db()
    rows = conn.execute("""SELECT c.name, c.level, c.capacity, COUNT(s.id),
        t.first_name||' '||t.last_name
        FROM classes c LEFT JOIN students s ON s.class_id=c.id AND s.status='actif'
        LEFT JOIN teachers t ON t.id=c.teacher_id WHERE c.cycle=? GROUP BY c.id""", (cycle,)).fetchall()
    conn.close()
    data = [[r[0],r[1] or '',r[2],r[3],r[4] or ''] for r in rows]
    return make_export(["Classe","Niveau","Capacite","Effectif","Titulaire"], data, f"Classes_{cycle.title()}_CongoSchool.xlsx")

@app.route('/export/subjects')
def export_subjects():
    conn = get_db()
    rows = conn.execute("""SELECT s.name, s.code, s.coefficient, t.first_name||' '||t.last_name, s.description
        FROM subjects s LEFT JOIN teachers t ON t.id=s.teacher_id ORDER BY s.name""").fetchall()
    conn.close()
    data = [[r[0],r[1] or '',r[2],r[3] or '',r[4] or ''] for r in rows]
    return make_export(["Matiere","Code","Coefficient","Enseignant","Description"], data, "Matieres_CongoSchool.xlsx")

@app.route('/export/grades')
def export_grades():
    cycle = session.get('cycle', 'primaire')
    conn = get_db()
    rows = conn.execute("""SELECT s.matricule, s.last_name||' '||s.first_name, c.name, sub.name,
        g.term, g.exam_type, g.score, g.max_score, ROUND(g.score*100.0/g.max_score,1), g.date
        FROM grades g JOIN students s ON s.id=g.student_id LEFT JOIN classes c ON c.id=s.class_id
        JOIN subjects sub ON sub.id=g.subject_id WHERE s.cycle=? ORDER BY c.name, s.last_name, sub.name""", (cycle,)).fetchall()
    conn.close()
    data = [[r[0],r[1],r[2] or '',r[3],r[4],r[5],r[6],r[7],r[8],r[9]] for r in rows]
    return make_export(["Matricule","Eleve","Classe","Matiere","Trimestre","Type examen","Note","Maximum","Pourcentage","Date"], data, f"Notes_{cycle.title()}_CongoSchool.xlsx")

@app.route('/export/attendance')
def export_attendance():
    cycle = session.get('cycle', 'primaire')
    conn = get_db()
    f_date = request.args.get('date', date.today().isoformat())
    rows = conn.execute("""SELECT s.matricule, s.last_name||' '||s.first_name, c.name, a.status, a.comment
        FROM attendance a JOIN students s ON s.id=a.student_id LEFT JOIN classes c ON c.id=s.class_id
        WHERE a.date=? AND s.cycle=? ORDER BY c.name, s.last_name""", (f_date, cycle)).fetchall()
    conn.close()
    data = [[r[0],r[1],r[2] or '',r[3],r[4] or ''] for r in rows]
    return make_export(["Matricule","Eleve","Classe","Statut","Commentaire"], data, f"Presences_{cycle.title()}_{f_date}.xlsx")

@app.route('/export/timetable')
def export_timetable():
    cycle = session.get('cycle', 'primaire')
    conn = get_db()
    rows = conn.execute("""SELECT c.name, tt.day, tt.time_start, tt.time_end, sub.name, tt.room,
        t.first_name||' '||t.last_name
        FROM timetable tt JOIN classes c ON c.id=tt.class_id JOIN subjects sub ON sub.id=tt.subject_id
        LEFT JOIN teachers t ON t.id=sub.teacher_id
        WHERE c.cycle=?
        ORDER BY c.name, CASE tt.day WHEN 'Lundi' THEN 1 WHEN 'Mardi' THEN 2 WHEN 'Mercredi' THEN 3 WHEN 'Jeudi' THEN 4 WHEN 'Vendredi' THEN 5 WHEN 'Samedi' THEN 6 END, tt.time_start""", (cycle,)).fetchall()
    conn.close()
    data = [[r[0],r[1],r[2],r[3],r[4],r[5] or '',r[6] or ''] for r in rows]
    return make_export(["Classe","Jour","Debut","Fin","Matiere","Salle","Enseignant"], data, f"EmploiDuTemps_{cycle.title()}_CongoSchool.xlsx")

@app.route('/export/fees')
def export_fees():
    cycle = session.get('cycle', 'primaire')
    conn = get_db()
    rows = conn.execute("""SELECT s.matricule, s.last_name||' '||s.first_name, c.name, f.fee_type,
        f.amount, f.amount_paid, f.amount-f.amount_paid, f.due_date, f.payment_date,
        f.receipt_number, f.status
        FROM fees f JOIN students s ON s.id=f.student_id LEFT JOIN classes c ON c.id=s.class_id
        WHERE s.cycle=? ORDER BY s.last_name""", (cycle,)).fetchall()
    conn.close()
    data = [[r[0],r[1],r[2] or '',r[3],r[4],r[5],r[6],r[7] or '',r[8] or '',r[9] or '',r[10]] for r in rows]
    return make_export(["Matricule","Eleve","Classe","Type frais","Montant","Paye","Reste","Echeance","Date paiement","N Reçu","Statut"], data, f"FraisScolaires_{cycle.title()}_CongoSchool.xlsx")


# ──────────────────────────────────────────────
# CLOSE SCHOOL YEAR (Archive + Reset)
# ──────────────────────────────────────────────
def _export_all_to_zip():
    """Exporte toutes les données dans un fichier ZIP en mémoire."""
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Élèves
        conn = get_db()
        rows = conn.execute("""SELECT s.matricule, s.first_name, s.last_name, s.date_of_birth, s.gender,
            c.name, s.parent_name, s.parent_phone, s.address, s.email, s.enrollment_date, s.status
            FROM students s LEFT JOIN classes c ON c.id=s.class_id ORDER BY s.last_name""").fetchall()
        conn.close()
        _add_sheet_to_zip(zf, "Eleves", ["Matricule","Prénom","Nom","Date naissance","Sexe","Classe","Parent","Tél parent","Adresse","Email","Date inscription","Statut"], rows)

        # Enseignants
        conn = get_db()
        rows = conn.execute("SELECT first_name, last_name, phone, email, subject, address, hire_date, salary, status FROM teachers ORDER BY last_name").fetchall()
        conn.close()
        _add_sheet_to_zip(zf, "Enseignants", ["Prénom","Nom","Téléphone","Email","Matière","Adresse","Date embauche","Salaire","Statut"], rows)

        # Classes
        conn = get_db()
        rows = conn.execute("""SELECT c.name, c.level, c.capacity, COUNT(s.id),
            (SELECT first_name||' '||last_name FROM teachers WHERE id=c.teacher_id)
            FROM classes c LEFT JOIN students s ON s.class_id=c.id GROUP BY c.id ORDER BY c.name""").fetchall()
        conn.close()
        _add_sheet_to_zip(zf, "Classes", ["Classe","Niveau","Capacité","Effectif","Titulaire"], rows)

        # Matières
        conn = get_db()
        rows = conn.execute("""SELECT sb.name, sb.code, sb.coefficient,
            (SELECT first_name||' '||last_name FROM teachers WHERE id=sb.teacher_id), sb.description
            FROM subjects sb ORDER BY sb.name""").fetchall()
        conn.close()
        _add_sheet_to_zip(zf, "Matieres", ["Matière","Code","Coefficient","Enseignant","Description"], rows)

        # Notes
        conn = get_db()
        rows = conn.execute("""SELECT s.matricule, s.last_name||' '||s.first_name, c.name, sb.name,
            g.term, g.exam_type, g.score, g.max_score,
            ROUND(g.score*100.0/g.max_score,1) if g.max_score>0 else 0, g.date
            FROM grades g JOIN students s ON s.id=g.student_id
            LEFT JOIN classes c ON c.id=s.class_id LEFT JOIN subjects sb ON sb.id=g.subject_id
            ORDER BY s.last_name""").fetchall()
        conn.close()
        _add_sheet_to_zip(zf, "Notes", ["Matricule","Élève","Classe","Matière","Trimestre","Type examen","Note","Maximum","Pourcentage","Date"], rows)

        # Présences
        conn = get_db()
        rows = conn.execute("""SELECT s.matricule, s.last_name||' '||s.first_name, c.name, a.status, a.comment
            FROM attendance a JOIN students s ON s.id=a.student_id LEFT JOIN classes c ON c.id=s.class_id
            ORDER BY a.date DESC""").fetchall()
        conn.close()
        _add_sheet_to_zip(zf, "Presences", ["Matricule","Élève","Classe","Statut","Commentaire"], rows)

        # Emploi du temps
        conn = get_db()
        rows = conn.execute("""SELECT c.name, t.day_of_week, t.start_time, t.end_time, sb.name, t.room,
            (SELECT first_name||' '||last_name FROM teachers WHERE id=t.teacher_id)
            FROM timetable t LEFT JOIN classes c ON c.id=t.class_id
            LEFT JOIN subjects sb ON sb.id=t.subject_id ORDER BY t.day_of_week, t.start_time""").fetchall()
        conn.close()
        _add_sheet_to_zip(zf, "EmploiDuTemps", ["Classe","Jour","Début","Fin","Matière","Salle","Enseignant"], rows)

        # Frais scolaires
        conn = get_db()
        rows = conn.execute("""SELECT s.matricule, s.last_name||' '||s.first_name, c.name, f.fee_type,
            f.amount, f.amount_paid, f.amount-f.amount_paid, f.due_date, f.payment_date,
            f.receipt_number, f.status
            FROM fees f JOIN students s ON s.id=f.student_id LEFT JOIN classes c ON c.id=s.class_id
            ORDER BY s.last_name""").fetchall()
        conn.close()
        _add_sheet_to_zip(zf, "FraisScolaires", ["Matricule","Élève","Classe","Type frais","Montant","Payé","Reste","Échéance","Date paiement","N° Reçu","Statut"], rows)

    zip_buf.seek(0)
    return zip_buf

def _add_sheet_to_zip(zf, sheet_name, headers, rows):
    """Ajoute un fichier Excel dans le ZIP."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=i, column=i, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val if val is not None else "")
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18
    xls_buf = BytesIO()
    wb.save(xls_buf)
    xls_buf.seek(0)
    zf.writestr(f"{sheet_name}.xlsx", xls_buf.getvalue())

@app.route('/close_year', methods=['POST'])
def close_year():
    confirm = request.form.get('confirm_text', '').strip()
    if confirm != 'CONGO':
        flash("⚠️ Veuillez taper CONGO pour confirmer la clôture.", "danger")
        return redirect(url_for('dashboard'))

    cycle = session.get('cycle', 'primaire')

    # 1. Exporter tout en ZIP
    zip_buf = _export_all_to_zip()
    zip_filename = f"CongoSchool_Archive_{cycle.title()}_{date.today().strftime('%Y-%m-%d')}.zip"

    # 2. Supprimer les données du cycle actif (garder profs, classes, matières)
    conn = get_db()
    # Get student ids for this cycle
    student_ids = [r[0] for r in conn.execute("SELECT id FROM students WHERE cycle=?", (cycle,)).fetchall()]
    if student_ids:
        placeholders = ",".join("?" * len(student_ids))
        conn.execute(f"DELETE FROM grades WHERE student_id IN ({placeholders})", student_ids)
        conn.execute(f"DELETE FROM attendance WHERE student_id IN ({placeholders})", student_ids)
        conn.execute(f"DELETE FROM fees WHERE student_id IN ({placeholders})", student_ids)
        conn.execute(f"DELETE FROM students WHERE cycle=?", (cycle,))
        # Delete timetable entries for classes of this cycle
        class_ids = [r[0] for r in conn.execute("SELECT id FROM classes WHERE cycle=?", (cycle,)).fetchall()]
        if class_ids:
            class_placeholders = ",".join("?" * len(class_ids))
            conn.execute(f"DELETE FROM timetable WHERE class_id IN ({class_placeholders})", class_ids)
    conn.commit()
    conn.close()

    # 3. Envoyer le ZIP
    return send_file(zip_buf, as_attachment=True, download_name=zip_filename,
                     mimetype='application/zip')


# ──────────────────────────────────────────────
# START
# ──────────────────────────────────────────────
if __name__ == '__main__':
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('8.8.8.8', 80))
        local_ip = s.getsockname()[0]
    except Exception:
        local_ip = '127.0.0.1'
    finally:
        s.close()
    print(f'\n  Ouverture de CongoSchool...')
    print(f'  Ordinateur : http://localhost:5000')
    print(f'  Telephone  : http://{local_ip}:5000')
    print(f'  (Assurez-vous d\'etre sur le meme WiFi)')
    print()
    app.run(host='0.0.0.0', port=5000, debug=False)
